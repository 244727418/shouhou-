import sys
import sqlite3
import re
from datetime import datetime, timedelta

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QGroupBox, QLabel, QComboBox, QLineEdit, QCheckBox, QPushButton, QTableWidget,
    QTableWidgetItem, QMessageBox, QFileDialog, QInputDialog, QHeaderView, QAbstractItemView,
    QFrame, QStatusBar, QDateEdit, QDialog, QDialogButtonBox, QFormLayout, QShortcut, QAction, QMenu,
    QColorDialog, QListWidget, QListWidgetItem, QItemDelegate, QFontDialog, QSpinBox, QSlider, QSplitter,
    QSizePolicy
)
from PyQt5.QtCore import Qt, QDate, pyqtSignal, QTimer, QRect, QPoint
from PyQt5.QtGui import QColor, QKeySequence, QClipboard, QFont, QPalette

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import xlrd  # 用于支持 .xls 文件，需要安装 xlrd


# 自定义多选下拉框组件（基于复选框）
class MultiSelectComboBox(QWidget):
    itemsChanged = pyqtSignal()  # 定义信号
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.selected_items = set()
        self.items = []
        self.init_ui()

    def init_ui(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # 下拉按钮
        self.dropdown_btn = QPushButton("选择退款原因 ▼")
        self.dropdown_btn.setFixedSize(150, 30)
        self.dropdown_btn.setStyleSheet("""
            QPushButton { 
                border: 1px solid #ccc; 
                border-radius: 3px; 
                padding: 5px; 
                text-align: left; 
                background-color: white;
            }
            QPushButton:hover {
                background-color: #f0f0f0;
            }
        """)
        self.dropdown_btn.clicked.connect(self.toggle_dropdown)
        layout.addWidget(self.dropdown_btn)
        
        # 下拉窗口
        self.dropdown_widget = QWidget()
        self.dropdown_widget.setWindowFlags(Qt.Popup)
        self.dropdown_widget.setFixedSize(300, 200)
        self.dropdown_widget.setStyleSheet("""
            QWidget {
                border: 1px solid #ccc;
                border-radius: 3px;
                background-color: white;
            }
        """)
        
        dropdown_layout = QVBoxLayout(self.dropdown_widget)
        
        # 搜索框
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("搜索退款原因...")
        self.search_edit.textChanged.connect(self.filter_items)
        dropdown_layout.addWidget(self.search_edit)
        
        # 全选/清空按钮
        button_layout = QHBoxLayout()
        self.select_all_btn = QPushButton("全选")
        self.select_all_btn.clicked.connect(self.select_all)
        self.clear_btn = QPushButton("清空")
        self.clear_btn.clicked.connect(self.clear_selection)
        button_layout.addWidget(self.select_all_btn)
        button_layout.addWidget(self.clear_btn)
        dropdown_layout.addLayout(button_layout)
        
        # 选项列表（使用QListWidget + 复选框）
        self.list_widget = QListWidget()
        self.list_widget.setSelectionMode(QListWidget.NoSelection)  # 禁用选择，使用复选框
        dropdown_layout.addWidget(self.list_widget)
        
        self.dropdown_widget.hide()

    def addItems(self, items):
        """添加选项"""
        self.items = items
        self.update_list_widget()

    def update_list_widget(self):
        """更新列表控件"""
        self.list_widget.clear()
        
        for item in self.items:
            list_item = QListWidgetItem(item)
            list_item.setFlags(list_item.flags() | Qt.ItemIsUserCheckable)
            list_item.setCheckState(Qt.Unchecked)
            self.list_widget.addItem(list_item)
        
        # 连接复选框状态变化信号
        self.list_widget.itemChanged.connect(self.on_item_changed)

    def filter_items(self, text):
        """过滤选项"""
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            item.setHidden(text.lower() not in item.text().lower())

    def checkedItems(self):
        """获取选中的项目"""
        checked = []
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            if item.checkState() == Qt.Checked:
                checked.append(item.text())
        return checked

    def clearChecked(self):
        """清空选择"""
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            item.setCheckState(Qt.Unchecked)

    def select_all(self):
        """全选"""
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            if not item.isHidden():
                item.setCheckState(Qt.Checked)

    def clear_selection(self):
        """清空选择"""
        self.clearChecked()

    def on_item_changed(self, item):
        """复选框状态变化处理"""
        self.update_display()

    def toggle_dropdown(self):
        """切换下拉列表显示"""
        if self.dropdown_widget.isVisible():
            self.dropdown_widget.hide()
        else:
            # 显示在下拉按钮下方
            pos = self.dropdown_btn.mapToGlobal(QPoint(0, self.dropdown_btn.height()))
            self.dropdown_widget.move(pos)
            self.dropdown_widget.show()
            self.search_edit.setFocus()

    def update_display(self):
        """更新按钮显示"""
        selected = self.checkedItems()
        if selected:
            # 显示已选项数量
            if len(selected) == 1:
                self.dropdown_btn.setText(f"{selected[0]} ▼")
            else:
                self.dropdown_btn.setText(f"已选{len(selected)}项 ▼")
        else:
            self.dropdown_btn.setText("选择退款原因 ▼")
        
        # 触发变化信号
        self.itemsChanged.emit()

    def setMaximumWidth(self, width):
        """设置最大宽度"""
        self.dropdown_btn.setMaximumWidth(width)
        self.setFixedWidth(width)


# ---------------------------- 店铺基本信息设置对话框 --------------------------------
class StoreSettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("店铺基本信息设置")
        self.setFixedSize(400, 300)
        
        layout = QVBoxLayout()
        
        # 表单布局
        form_layout = QFormLayout()
        
        # 日单量
        self.daily_orders_edit = QLineEdit()
        self.daily_orders_edit.setPlaceholderText("请输入日单量")
        form_layout.addRow("日单量：", self.daily_orders_edit)
        
        # 日销售金额
        self.daily_sales_edit = QLineEdit()
        self.daily_sales_edit.setPlaceholderText("请输入日销售金额")
        form_layout.addRow("日销售金额：", self.daily_sales_edit)
        
        # 退款预算设置
        refund_budget_layout = QHBoxLayout()
        
        self.refund_budget_amount_edit = QLineEdit()
        self.refund_budget_amount_edit.setPlaceholderText("输入金额")
        self.refund_budget_amount_edit.textChanged.connect(self.on_amount_changed)
        refund_budget_layout.addWidget(self.refund_budget_amount_edit)
        
        refund_budget_layout.addWidget(QLabel("或"))
        
        self.refund_budget_percent_edit = QLineEdit()
        self.refund_budget_percent_edit.setPlaceholderText("输入百分比")
        self.refund_budget_percent_edit.textChanged.connect(self.on_percent_changed)
        refund_budget_layout.addWidget(self.refund_budget_percent_edit)
        
        refund_budget_layout.addWidget(QLabel("%"))
        
        form_layout.addRow("退款预算：", refund_budget_layout)
        
        layout.addLayout(form_layout)
        layout.addStretch()
        
        # 按钮布局
        button_layout = QHBoxLayout()
        
        self.save_btn = QPushButton("保存")
        self.save_btn.clicked.connect(self.save_settings)
        button_layout.addWidget(self.save_btn)
        
        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)

    def on_amount_changed(self, text):
        """金额输入变化时自动计算百分比"""
        if text and self.daily_sales_edit.text():
            try:
                amount = float(text)
                sales = float(self.daily_sales_edit.text())
                if sales > 0:
                    percent = (amount / sales) * 100
                    # 临时断开信号避免循环
                    self.refund_budget_percent_edit.textChanged.disconnect(self.on_percent_changed)
                    self.refund_budget_percent_edit.setText(f"{percent:.2f}")
                    self.refund_budget_percent_edit.textChanged.connect(self.on_percent_changed)
            except ValueError:
                pass

    def on_percent_changed(self, text):
        """百分比输入变化时自动计算金额"""
        if text and self.daily_sales_edit.text():
            try:
                percent = float(text)
                sales = float(self.daily_sales_edit.text())
                amount = (percent / 100) * sales
                # 临时断开信号避免循环
                self.refund_budget_amount_edit.textChanged.disconnect(self.on_amount_changed)
                self.refund_budget_amount_edit.setText(f"{amount:.2f}")
                self.refund_budget_amount_edit.textChanged.connect(self.on_amount_changed)
            except ValueError:
                pass

    def save_settings(self):
        """保存设置"""
        try:
            # 验证输入
            daily_orders = int(self.daily_orders_edit.text()) if self.daily_orders_edit.text() else 0
            daily_sales = float(self.daily_sales_edit.text()) if self.daily_sales_edit.text() else 0.0
            
            # 获取退款预算（优先使用金额输入）
            if self.refund_budget_amount_edit.text():
                refund_budget = float(self.refund_budget_amount_edit.text())
            elif self.refund_budget_percent_edit.text():
                percent = float(self.refund_budget_percent_edit.text())
                refund_budget = (percent / 100) * daily_sales
            else:
                refund_budget = 0.0
            
            # 保存到主窗口
            self.parent.store_settings = {
                'daily_orders': daily_orders,
                'daily_sales': daily_sales,
                'refund_budget': refund_budget
            }
            
            # 保存到数据库
            current_store = self.parent.search_store_combo.currentText()
            
            if current_store and current_store != "全部":
                # 保存到当前店铺
                stores = self.parent.db.get_stores()
                store_id = None
                for sid, sname in stores:
                    if sname == current_store:
                        store_id = sid
                        break
                
                if store_id:
                    # 更新店铺设置到数据库
                    self.parent.db.update_store_settings(store_id, daily_orders, daily_sales, refund_budget)
            else:
                # 选择"全部"店铺时，保存到全局设置
                self.parent.db.save_global_settings(daily_orders, daily_sales, refund_budget)
            
            # 更新显示
            self.parent.update_store_stats_display()
            
            self.accept()
            
        except ValueError:
            QMessageBox.warning(self, "输入错误", "请输入有效的数字")

    def load_settings(self, settings):
        """加载现有设置"""
        # 优先从数据库加载设置
        current_store = self.parent.search_store_combo.currentText()
        if current_store and current_store != "全部":
            # 获取店铺ID
            stores = self.parent.db.get_stores()
            store_id = None
            for sid, sname in stores:
                if sname == current_store:
                    store_id = sid
                    break
            
            if store_id:
                # 从数据库加载设置
                db_settings = self.parent.db.get_store_settings(store_id)
                if db_settings:
                    settings = db_settings
        
        if settings:
            self.daily_orders_edit.setText(str(settings.get('daily_orders', 0)))
            self.daily_sales_edit.setText(str(settings.get('daily_sales', 0.0)))
            
            refund_budget = settings.get('refund_budget', 0.0)
            self.refund_budget_amount_edit.setText(str(refund_budget))
            
            # 自动计算百分比
            if settings.get('daily_sales', 0.0) > 0:
                percent = (refund_budget / settings['daily_sales']) * 100
                self.refund_budget_percent_edit.setText(f"{percent:.2f}")


# ---------------------------- 自定义表格委托类 --------------------------------
class CustomItemDelegate(QItemDelegate):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent

    def createEditor(self, parent, option, index):
        """创建编辑器时检查编辑权限"""
        # 补偿金额列（第6列）检查打款补偿状态
        if index.column() == 6:  # 补偿金额列
            # 获取当前行的记录信息
            row = index.row()
            record_id = self.parent.get_record_id_from_row(row)
            if record_id:
                record = self.parent.db.get_record_by_id(record_id)
                if record and not record['compensate']:
                    # 如果没有勾选打款补偿，不允许编辑
                    return None
        
        # 其他列正常创建编辑器
        return super().createEditor(parent, option, index)
    
    def setEditorData(self, editor, index):
        """设置编辑器数据，在编辑时保持选中状态"""
        # 在开始编辑时，确保当前行保持选中状态
        if self.parent and hasattr(self.parent, 'table'):
            # 获取当前行
            row = index.row()
            # 确保该行被选中
            self.parent.table.setCurrentCell(row, index.column())
        
        # 调用父类方法设置编辑器数据
        super().setEditorData(editor, index)
    
    def setModelData(self, editor, model, index):
        """设置模型数据，在编辑完成后保持选中状态"""
        # 调用父类方法设置模型数据
        super().setModelData(editor, model, index)
        
        # 在编辑完成后，确保当前行保持选中状态
        if self.parent and hasattr(self.parent, 'table'):
            # 获取当前行
            row = index.row()
            # 确保该行被选中
            self.parent.table.setCurrentCell(row, index.column())

# ---------------------------- 数据库操作类 ---------------------------------
class Database:
    def __init__(self, db_file='refund_data.db'):
        self.db_file = db_file
        self.conn = None
        self.init_db()

    def init_db(self):
        """初始化数据库，创建表"""
        self.conn = sqlite3.connect(self.db_file)
        # 启用外键约束（SQLite默认不启用）
        self.conn.execute("PRAGMA foreign_keys = ON")
        cursor = self.conn.cursor()
        
        # 检查表是否存在，如果存在则添加缺失的列
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='stores'")
        if cursor.fetchone():
            # 表已存在，检查并添加缺失的列
            self._add_missing_columns()
        else:
            # 表不存在，创建新表
            cursor.execute('''
                CREATE TABLE stores (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    store_name TEXT UNIQUE NOT NULL,
                    color TEXT DEFAULT NULL,
                    estimated_orders INTEGER DEFAULT 0,
                    daily_orders INTEGER DEFAULT 0,
                    daily_sales REAL DEFAULT 0.0,
                    refund_budget REAL DEFAULT 0.0
                )
            ''')
        
        # 性能优化：自动修复缺失的表（确保exe运行时表一定存在）
        self._auto_fix_missing_tables()
        
        # 创建 refund_records 表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS refund_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                store_id INTEGER NOT NULL,
                order_no TEXT NOT NULL,
                reason TEXT NOT NULL,
                refund_amount REAL NOT NULL,
                cancel INTEGER DEFAULT 0,
                compensate INTEGER DEFAULT 0,
                comp_amount REAL DEFAULT 0,
                reject INTEGER DEFAULT 0,  -- 是否驳回：0=否，1=是
                reject_result TEXT DEFAULT '',  -- 驳回结果：成功、失败
                notes TEXT DEFAULT '',  -- 备注信息
                record_date TEXT DEFAULT '',
                FOREIGN KEY (store_id) REFERENCES stores (id) ON DELETE CASCADE
            )
        ''')
        # 添加索引
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_order_no ON refund_records (order_no)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_record_date ON refund_records (record_date)')
        self.conn.commit()
    
    def _add_missing_columns(self):
        """添加缺失的列到stores表"""
        cursor = self.conn.cursor()
        
        # 检查daily_orders列是否存在
        cursor.execute("PRAGMA table_info(stores)")
        columns = [column[1] for column in cursor.fetchall()]
        
        if 'daily_orders' not in columns:
            cursor.execute("ALTER TABLE stores ADD COLUMN daily_orders INTEGER DEFAULT 0")
        
        if 'daily_sales' not in columns:
            cursor.execute("ALTER TABLE stores ADD COLUMN daily_sales REAL DEFAULT 0.0")
        
        if 'refund_budget' not in columns:
            cursor.execute("ALTER TABLE stores ADD COLUMN refund_budget REAL DEFAULT 0.0")
        
        # 创建全局设置表（用于存储"全部店铺"的设置）
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS global_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                setting_key TEXT UNIQUE NOT NULL,
                setting_value TEXT
            )
        ''')
        
        # 创建窗口设置表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS window_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                setting_key TEXT UNIQUE NOT NULL,
                setting_value TEXT
            )
        ''')
        
        self.conn.commit()

    def _auto_fix_missing_tables(self):
        """自动修复缺失的表（确保exe运行时表一定存在）"""
        cursor = self.conn.cursor()
        
        # 检查 global_settings 表是否存在
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='global_settings'")
        if not cursor.fetchone():
            # 创建 global_settings 表
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS global_settings (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    setting_key TEXT UNIQUE NOT NULL,
                    setting_value TEXT
                )
            ''')
            print("✅ 自动修复：global_settings 表已创建")
        
        # 检查 window_settings 表是否存在
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='window_settings'")
        if not cursor.fetchone():
            # 创建 window_settings 表
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS window_settings (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    setting_key TEXT UNIQUE NOT NULL,
                    setting_value TEXT
                )
            ''')
            print("✅ 自动修复：window_settings 表已创建")
        
        # 检查 refund_records 表是否存在
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='refund_records'")
        if not cursor.fetchone():
            # 创建 refund_records 表
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS refund_records (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    store_id INTEGER NOT NULL,
                    order_no TEXT NOT NULL,
                    reason TEXT NOT NULL,
                    refund_amount REAL NOT NULL,
                    cancel INTEGER DEFAULT 0,
                    compensate INTEGER DEFAULT 0,
                    comp_amount REAL DEFAULT 0,
                    reject INTEGER DEFAULT 0,
                    reject_result TEXT DEFAULT '',
                    notes TEXT DEFAULT '',
                    record_date TEXT DEFAULT '',
                    FOREIGN KEY (store_id) REFERENCES stores (id) ON DELETE CASCADE
                )
            ''')
            print("✅ 自动修复：refund_records 表已创建")
        
        self.conn.commit()

    def close(self):
        """关闭数据库连接"""
        if self.conn:
            self.conn.close()

    def get_stores(self):
        """获取所有店铺，返回列表 [(id, name), ...]"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT id, store_name FROM stores ORDER BY store_name')
        return cursor.fetchall()

    def add_store(self, name):
        """添加店铺，返回新ID，如果已存在返回None"""
        try:
            cursor = self.conn.cursor()
            cursor.execute('INSERT INTO stores (store_name) VALUES (?)', (name,))
            self.conn.commit()
            return cursor.lastrowid
        except sqlite3.IntegrityError:
            return None

    def set_store_color(self, store_name, color):
        """设置店铺颜色"""
        cursor = self.conn.cursor()
        cursor.execute('UPDATE stores SET color = ? WHERE store_name = ?', (color, store_name))
        self.conn.commit()
        return cursor.rowcount > 0

    def get_store_color(self, store_name):
        """获取店铺颜色"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT color FROM stores WHERE store_name = ?', (store_name,))
        result = cursor.fetchone()
        return result[0] if result and result[0] else None

    def clear_store_color(self, store_name):
        """清除店铺颜色"""
        cursor = self.conn.cursor()
        cursor.execute('UPDATE stores SET color = NULL WHERE store_name = ?', (store_name,))
        self.conn.commit()
        return cursor.rowcount > 0

    def set_estimated_orders(self, store_name, estimated_orders):
        """设置店铺预估订单量"""
        cursor = self.conn.cursor()
        cursor.execute('UPDATE stores SET estimated_orders = ? WHERE store_name = ?', (estimated_orders, store_name))
        self.conn.commit()
        return cursor.rowcount > 0

    def get_estimated_orders(self, store_name):
        """获取店铺预估订单量"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT estimated_orders FROM stores WHERE store_name = ?', (store_name,))
        result = cursor.fetchone()
        return result[0] if result else 0

    def update_store_settings(self, store_id, daily_orders, daily_sales, refund_budget):
        """更新店铺设置"""
        cursor = self.conn.cursor()
        cursor.execute('''
            UPDATE stores SET 
                daily_orders = ?, daily_sales = ?, refund_budget = ?
            WHERE id = ?
        ''', (daily_orders, daily_sales, refund_budget, store_id))
        self.conn.commit()
        return cursor.rowcount > 0

    def get_store_settings(self, store_id):
        """获取店铺设置"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT daily_orders, daily_sales, refund_budget
            FROM stores WHERE id = ?
        ''', (store_id,))
        result = cursor.fetchone()
        if result:
            return {
                'daily_orders': result[0],
                'daily_sales': result[1],
                'refund_budget': result[2]
            }
        return None

    def delete_store(self, store_id):
        """删除店铺及其相关数据（由于外键约束，相关记录会自动删除）"""
        try:
            cursor = self.conn.cursor()
            cursor.execute('DELETE FROM stores WHERE id = ?', (store_id,))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"删除店铺失败: {e}")
            return False

    def update_store_name(self, store_id, new_name):
        """修改店铺名称"""
        try:
            cursor = self.conn.cursor()
            cursor.execute('UPDATE stores SET store_name = ? WHERE id = ?', (new_name, store_id))
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            # 店铺名称已存在
            return False
        except Exception as e:
            print(f"修改店铺名称失败: {e}")
            return False

    def save_global_settings(self, daily_orders, daily_sales, refund_budget):
        """保存全局设置（全部店铺）"""
        cursor = self.conn.cursor()
        
        # 保存设置到全局设置表
        settings = {
            'daily_orders': daily_orders,
            'daily_sales': daily_sales,
            'refund_budget': refund_budget
        }
        
        for key, value in settings.items():
            cursor.execute('''
                INSERT OR REPLACE INTO global_settings (setting_key, setting_value)
                VALUES (?, ?)
            ''', (key, str(value)))
        
        self.conn.commit()

    def get_global_settings(self):
        """获取全局设置（全部店铺）"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT setting_key, setting_value FROM global_settings')
        results = cursor.fetchall()
        
        if not results:
            return {'daily_orders': 0, 'daily_sales': 0.0, 'refund_budget': 0.0}
        
        settings = {}
        for key, value in results:
            if key == 'daily_orders':
                settings[key] = int(value) if value else 0
            elif key in ['daily_sales', 'refund_budget']:
                settings[key] = float(value) if value else 0.0
            else:
                settings[key] = value
        
        return settings

    def save_window_settings(self, settings):
        """保存窗口设置到数据库"""
        cursor = self.conn.cursor()
        
        for key, value in settings.items():
            cursor.execute('''
                INSERT OR REPLACE INTO window_settings (setting_key, setting_value)
                VALUES (?, ?)
            ''', (key, str(value)))
        
        self.conn.commit()

    def load_window_settings(self):
        """从数据库加载窗口设置"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT setting_key, setting_value FROM window_settings')
        results = cursor.fetchall()
        
        if not results:
            return None
        
        settings = {}
        for key, value in results:
            # 解析不同类型的设置值
            if key in ['window_size', 'main_splitter', 'top_splitter', 'bottom_splitter']:
                # 列表类型设置（如分割器比例）
                try:
                    settings[key] = eval(value)  # 使用eval将字符串转换为列表
                except:
                    settings[key] = []
            elif value.isdigit():
                settings[key] = int(value)
            elif value.replace('.', '', 1).isdigit():
                settings[key] = float(value)
            else:
                settings[key] = value
        
        return settings

    def get_store_refund_stats(self, store_name):
        """获取店铺退款统计（排除撤销订单）"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT COUNT(*), SUM(refund_amount), SUM(comp_amount)
            FROM refund_records r
            JOIN stores s ON r.store_id = s.id
            WHERE s.store_name = ? AND r.cancel = 0
        ''', (store_name,))
        result = cursor.fetchone()
        if result and result[0] is not None:
            return {
                'refund_count': result[0],
                'total_refund': result[1] or 0.0,
                'total_comp': result[2] or 0.0
            }
        return {'refund_count': 0, 'total_refund': 0.0, 'total_comp': 0.0}

    def get_all_records(self):
        """获取所有退款记录"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT r.id, r.order_no, r.reason, r.refund_amount, r.cancel, r.compensate, r.comp_amount, r.record_date, s.store_name, r.store_id
            FROM refund_records r
            JOIN stores s ON r.store_id = s.id
            ORDER BY r.record_date DESC, r.id DESC
        ''')
        records = []
        for row in cursor.fetchall():
            records.append({
                'id': row[0], 'order_no': row[1], 'reason': row[2], 'refund_amount': row[3],
                'cancel': bool(row[4]), 'compensate': bool(row[5]), 'comp_amount': row[6],
                'record_date': row[7], 'store_name': row[8], 'store_id': row[9]
            })
        return records

    def get_total_record_count(self):
        """获取数据库中的总记录数"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM refund_records')
        result = cursor.fetchone()
        if result and isinstance(result, (tuple, list)) and len(result) > 0:
            return result[0] if isinstance(result[0], int) else int(result[0])
        return 0

    def cleanup_orphan_records(self):
        """清理没有对应店铺的孤儿记录"""
        try:
            cursor = self.conn.cursor()
            # 删除没有对应店铺的记录
            cursor.execute('''
                DELETE FROM refund_records 
                WHERE store_id NOT IN (SELECT id FROM stores)
            ''')
            deleted_count = cursor.rowcount
            self.conn.commit()
            return deleted_count
        except Exception as e:
            print(f"清理孤儿记录失败: {e}")
            return 0

    def debug_database_records(self):
        """调试功能：查看数据库中的所有记录"""
        try:
            cursor = self.conn.cursor()
            cursor.execute('''
                SELECT r.id, r.order_no, r.store_id, s.store_name, r.reason, r.record_date
                FROM refund_records r
                LEFT JOIN stores s ON r.store_id = s.id
                ORDER BY r.id
            ''')
            records = cursor.fetchall()
            
            result = []
            for record in records:
                result.append({
                    'id': record[0],
                    'order_no': record[1],
                    'store_id': record[2],
                    'store_name': record[3] if record[3] else '无对应店铺',
                    'reason': record[4],
                    'record_date': record[5]
                })
            
            return result
        except Exception as e:
            print(f"调试查询失败: {e}")
            return []

    def force_global_sync(self):
        """强制全局同步：彻底清理所有不一致数据"""
        try:
            cursor = self.conn.cursor()
            
            # 第一步：清理孤儿记录
            cursor.execute('''
                DELETE FROM refund_records 
                WHERE store_id NOT IN (SELECT id FROM stores)
            ''')
            orphan_count = cursor.rowcount
            
            # 第二步：清理重复记录（保留最新的）
            cursor.execute('''
                DELETE FROM refund_records 
                WHERE id NOT IN (
                    SELECT MAX(id) 
                    FROM refund_records 
                    GROUP BY order_no, store_id
                )
            ''')
            duplicate_count = cursor.rowcount
            
            # 第三步：清理无效数据（订单号为空或店铺ID为0）
            cursor.execute('''
                DELETE FROM refund_records 
                WHERE order_no = '' OR order_no IS NULL OR store_id = 0
            ''')
            invalid_count = cursor.rowcount
            
            # 第四步：清理所有隐藏的不一致数据（终极清理）
            cursor.execute('''
                DELETE FROM refund_records 
                WHERE id NOT IN (
                    SELECT r.id 
                    FROM refund_records r
                    JOIN stores s ON r.store_id = s.id
                )
            ''')
            hidden_count = cursor.rowcount
            
            self.conn.commit()
            
            return {
                'orphan_count': orphan_count,
                'duplicate_count': duplicate_count,
                'invalid_count': invalid_count,
                'hidden_count': hidden_count,
                'total_cleaned': orphan_count + duplicate_count + invalid_count + hidden_count
            }
        except Exception as e:
            print(f"强制同步失败: {e}")
            return {'orphan_count': 0, 'duplicate_count': 0, 'invalid_count': 0, 'hidden_count': 0, 'total_cleaned': 0}

    def get_filtered_record_count(self, order_no='', reason='全部', cancel='全部', compensate='全部',
                                 reject='全部', reject_result='全部', start_date=None, end_date=None, store_name='全部'):
        """根据筛选条件获取记录数"""
        cursor = self.conn.cursor()
        query = 'SELECT COUNT(*) FROM refund_records r JOIN stores s ON r.store_id = s.id WHERE 1=1'
        params = []
        
        if order_no:
            query += ' AND r.order_no LIKE ?'
            params.append(f'%{order_no}%')
        
        if reason != '全部':
            query += ' AND r.reason = ?'
            params.append(reason)
        
        if cancel != '全部':
            if cancel == '是':
                query += ' AND r.cancel = 1'
            elif cancel == '否':
                query += ' AND r.cancel = 0'
        
        if compensate != '全部':
            if compensate == '是':
                query += ' AND r.compensate = 1'
            elif compensate == '否':
                query += ' AND r.compensate = 0'
        
        if reject != '全部':
            if reject == '是':
                query += ' AND r.reject = 1'
            elif reject == '否':
                query += ' AND r.reject = 0'
        
        if reject_result != '全部':
            query += ' AND r.reject_result = ?'
            params.append(reject_result)
        
        if start_date:
            query += ' AND r.record_date >= ?'
            params.append(start_date)
        
        if end_date:
            query += ' AND r.record_date <= ?'
            params.append(end_date)
        
        if store_name != '全部':
            query += ' AND s.store_name = ?'
            params.append(store_name)
        
        cursor.execute(query, params)
        result = cursor.fetchone()
        if result and isinstance(result, (tuple, list)) and len(result) > 0:
            return result[0] if isinstance(result[0], int) else int(result[0])
        return 0

    def add_record(self, store_id, order_no, reason, refund_amount, cancel, compensate, comp_amount, reject, reject_result, notes, record_date):
        """添加退款记录"""
        cursor = self.conn.cursor()
        cursor.execute('''
            INSERT INTO refund_records 
            (store_id, order_no, reason, refund_amount, cancel, compensate, comp_amount, reject, reject_result, notes, record_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (store_id, order_no, reason, refund_amount, 1 if cancel else 0, 1 if compensate else 0, comp_amount, 1 if reject else 0, reject_result, notes, record_date))
        self.conn.commit()
        return cursor.lastrowid

    def update_record(self, record_id, store_id, order_no, reason, refund_amount, cancel, compensate, comp_amount, reject, reject_result, notes, record_date):
        """更新退款记录"""
        cursor = self.conn.cursor()
        cursor.execute('''
            UPDATE refund_records SET
                store_id=?, order_no=?, reason=?, refund_amount=?,
                cancel=?, compensate=?, comp_amount=?, reject=?, reject_result=?, notes=?, record_date=?
            WHERE id=?
        ''', (store_id, order_no, reason, refund_amount, 1 if cancel else 0, 1 if compensate else 0, comp_amount, 1 if reject else 0, reject_result, notes, record_date, record_id))
        self.conn.commit()

    def update_refund_amount(self, record_id, refund_amount):
        """更新退款金额"""
        cursor = self.conn.cursor()
        cursor.execute('UPDATE refund_records SET refund_amount=? WHERE id=?', (refund_amount, record_id))
        self.conn.commit()
        return cursor.rowcount > 0

    def update_comp_amount(self, record_id, comp_amount):
        """更新补偿金额"""
        cursor = self.conn.cursor()
        cursor.execute('UPDATE refund_records SET comp_amount=? WHERE id=?', (comp_amount, record_id))
        self.conn.commit()
        return cursor.rowcount > 0

    def delete_record(self, record_id):
        """删除退款记录，返回是否成功"""
        cursor = self.conn.cursor()
        cursor.execute('DELETE FROM refund_records WHERE id=?', (record_id,))
        self.conn.commit()
        return cursor.rowcount > 0  # 返回删除是否成功

    def get_record_by_id(self, record_id):
        """根据ID获取记录"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT r.id, r.order_no, r.reason, r.refund_amount, r.cancel, r.compensate, r.comp_amount, 
                   r.reject, r.reject_result, r.notes, r.record_date, s.store_name, r.store_id
            FROM refund_records r
            JOIN stores s ON r.store_id = s.id
            WHERE r.id=?
        ''', (record_id,))
        row = cursor.fetchone()
        if row:
            return {
                'id': row[0], 'order_no': row[1], 'reason': row[2], 'refund_amount': row[3],
                'cancel': bool(row[4]), 'compensate': bool(row[5]), 'comp_amount': row[6],
                'reject': bool(row[7]), 'reject_result': row[8], 'notes': row[9],
                'record_date': row[10], 'store_name': row[11], 'store_id': row[12]
            }
        return None

    def is_order_no_exists(self, order_no):
        """检查订单号是否已存在"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT id FROM refund_records WHERE order_no = ?', (order_no,))
        return cursor.fetchone() is not None

    def get_record_by_order_no(self, order_no):
        """根据订单号获取记录"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT r.id, r.order_no, r.reason, r.refund_amount, r.cancel, r.compensate, r.comp_amount, 
                   r.reject, r.reject_result, r.notes, r.record_date, s.store_name, r.store_id
            FROM refund_records r
            JOIN stores s ON r.store_id = s.id
            WHERE r.order_no=?
        ''', (order_no,))
        row = cursor.fetchone()
        if row:
            return {
                'id': row[0], 'order_no': row[1], 'reason': row[2], 'refund_amount': row[3],
                'cancel': bool(row[4]), 'compensate': bool(row[5]), 'comp_amount': row[6],
                'reject': bool(row[7]), 'reject_result': row[8], 'notes': row[9],
                'record_date': row[10], 'store_name': row[11], 'store_id': row[12]
            }
        return None

    def search_records(self, order_no='', reason='全部', cancel='全部', compensate='全部',
                       reject='全部', reject_result='全部', start_date=None, end_date=None, store_name='全部'):
        """根据条件搜索记录，返回结果列表"""
        cursor = self.conn.cursor()
        query = '''
            SELECT r.id, r.order_no, r.reason, r.refund_amount, r.cancel, r.compensate, r.comp_amount, 
                   r.reject, r.reject_result, r.notes, r.record_date, s.store_name
            FROM refund_records r
            JOIN stores s ON r.store_id = s.id
            WHERE 1=1
        '''
        params = []
        if order_no:
            # 智能模糊搜索：去除输入中的空格、换行符等特殊字符，支持部分匹配
            # 清理用户输入：去除空格、换行符、制表符等
            order_no_cleaned = re.sub(r'\s+', '', order_no)  # 去除所有空白字符
            
            # 如果清理后还有内容，进行模糊搜索
            if order_no_cleaned:
                # 同时清理数据库中的订单号进行匹配
                query += ' AND REPLACE(REPLACE(r.order_no, " ", ""), "\n", "") LIKE ?'
                params.append(f'%{order_no_cleaned}%')
        if reason != '全部':
            if isinstance(reason, list) and len(reason) > 0:
                # 多选情况：使用 IN 查询
                placeholders = ','.join(['?'] * len(reason))
                query += f' AND r.reason IN ({placeholders})'
                params.extend(reason)
            else:
                # 单选情况
                query += ' AND r.reason = ?'
                params.append(reason)
        if cancel != '全部':
            query += ' AND r.cancel = ?'
            params.append(1 if cancel == '是' else 0)
        if compensate != '全部':
            query += ' AND r.compensate = ?'
            params.append(1 if compensate == '是' else 0)
        if reject != '全部':
            query += ' AND r.reject = ?'
            params.append(1 if reject == '是' else 0)
        if reject_result != '全部':
            query += ' AND r.reject_result = ?'
            params.append(reject_result)
        if store_name != '全部':
            query += ' AND s.store_name = ?'
            params.append(store_name)
        if start_date:
            query += ' AND r.record_date >= ?'
            params.append(start_date)
        if end_date:
            query += ' AND r.record_date <= ?'
            params.append(end_date)
        query += ' ORDER BY r.record_date DESC, r.id DESC'
        cursor.execute(query, params)
        rows = cursor.fetchall()
        results = []
        for row in rows:
            results.append({
                'id': row[0], 'order_no': row[1], 'reason': row[2], 'refund_amount': row[3],
                'cancel': bool(row[4]), 'compensate': bool(row[5]), 'comp_amount': row[6],
                'reject': bool(row[7]), 'reject_result': row[8], 'notes': row[9],
                'record_date': row[10], 'store_name': row[11]
            })
        return results

# ---------------------------- 主窗口类 ---------------------------------
class RefundManager(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = Database()
        self.current_record_id = None  # 当前编辑的记录ID（用于更新）
        self.highlighted_orders = set()  # 刚导入需要高亮的订单号集合
        self.selected_reasons = set()  # 多选退款原因集合
        self.store_settings = {}  # 店铺基本信息设置
        
        # 性能优化：初始化定时器（避免重复创建）
        self._search_timer = QTimer()
        self._search_timer.setSingleShot(True)
        
        # 性能优化：数据缓存
        self._cached_records = None  # 缓存搜索结果
        self._last_search_params = None  # 上次搜索参数
        
        self.init_ui()
        # 初始化店铺设置
        self.load_store_settings()
        self.load_stores()
        self.load_table_data()
        self.setup_shortcuts()

    def init_ui(self):
        self.setWindowTitle("电商售后品质退款管理工具")
        # 【窗口默认尺寸设置】第451行 - 修改这里的数字来改变窗口默认大小
        self.resize(1400, 950)  # 窗口宽度1400像素，高度950像素
        self.setMinimumSize(0, 0)  # 设置窗口最小尺寸，允许适当缩小
        
        # 应用护眼配色样式表
        self.apply_stylesheet()

        # 中央控件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 主布局：垂直分割器（上下分割）
        main_splitter = QSplitter(Qt.Vertical)
        main_splitter.setChildrenCollapsible(False)  # 禁止折叠子部件
        main_splitter.setStretchFactor(0, 1)  # 上部区域可拉伸
        main_splitter.setStretchFactor(1, 1)  # 下部区域可拉伸
        main_layout = QVBoxLayout(central_widget)
        main_layout.addWidget(main_splitter)
        
        # 上部区域：水平分割器（左右分割）
        top_splitter = QSplitter(Qt.Horizontal)
        top_splitter.setChildrenCollapsible(False)  # 禁止折叠子部件
        top_splitter.setStretchFactor(0, 1)  # 左侧信息录入区可拉伸
        top_splitter.setStretchFactor(1, 1)  # 右侧店铺信息区可拉伸
        
        # 下部区域：水平分割器（左右分割）
        bottom_splitter = QSplitter(Qt.Horizontal)
        bottom_splitter.setChildrenCollapsible(False)  # 禁止折叠子部件
        bottom_splitter.setHandleWidth(8)  # 增加分割器手柄宽度，便于拖拽
        bottom_splitter.setStretchFactor(0, 1)  # 左侧搜索筛选区可拉伸
        bottom_splitter.setStretchFactor(1, 1)  # 右侧表格区可拉伸
        
        # 将上下分割器添加到主分割器
        main_splitter.addWidget(top_splitter)
        main_splitter.addWidget(bottom_splitter)
        
        # 设置分割器最小尺寸，防止折叠和瞬间变0
        main_splitter.setMinimumSize(800, 600)  # 主窗口最小尺寸
        top_splitter.setMinimumSize(0, 0)        # 上部区域最小尺寸：左右都为0，完全自由调整
        bottom_splitter.setMinimumSize(200, 300) # 下部区域最小尺寸：左200，右300
        
        # 保存分割器引用，用于记忆功能
        self.main_splitter = main_splitter
        self.top_splitter = top_splitter
        self.bottom_splitter = bottom_splitter
        
        # 左上角：信息录入区
        input_group = QGroupBox("信息录入")
        input_layout = QGridLayout()
        input_layout.setHorizontalSpacing(0)  # 设置列间距为0，让标签和输入框紧挨着
        input_layout.setVerticalSpacing(5)    # 保持行间距为5像素
        input_group.setLayout(input_layout)

        # 第一行：店铺、订单号、退款原因
        store_label = QLabel("店铺：")
        store_label.setStyleSheet("margin-right: 2px; padding: 0px; font-size: 14px;")
        input_layout.addWidget(store_label, 0, 0)
        
        # 店铺选择区域 - 使用水平布局包含下拉框和操作按钮
        store_widget = QWidget()
        store_layout = QHBoxLayout(store_widget)
        store_layout.setContentsMargins(0, 0, 0, 0)
        store_layout.setSpacing(5)
        
        # 美化后的店铺下拉框
        self.store_combo = QComboBox()
        self.store_combo.setStyleSheet("""
            QComboBox {
                font-size: 16px;
                font-weight: bold;
                padding: 8px 12px;
                border: 2px solid #4CAF50;
                border-radius: 6px;
                background-color: white;
                min-height: 35px;
                min-width: 200px;
            }
            QComboBox:hover {
                border-color: #45a049;
                background-color: #f8fff8;
            }
            QComboBox:focus {
                border-color: #2196F3;
                background-color: #f0f8ff;
            }
            QComboBox::drop-down {
                border: none;
                width: 30px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 8px solid #666;
                width: 0;
                height: 0;
            }
            QComboBox QAbstractItemView {
                font-size: 14px;
                border: 2px solid #4CAF50;
                border-radius: 6px;
                background-color: white;
                selection-background-color: #4CAF50;
                selection-color: white;
                outline: none;
            }
        """)
        self.store_combo.currentTextChanged.connect(self.on_store_combo_changed)  # 同步店铺选择
        store_layout.addWidget(self.store_combo)
        
        # 添加店铺按钮
        self.add_store_btn = QPushButton("➕")
        self.add_store_btn.setToolTip("添加新店铺")
        self.add_store_btn.setStyleSheet("""
            QPushButton {
                font-size: 16px;
                padding: 8px 12px;
                border: 2px solid #2196F3;
                border-radius: 6px;
                background-color: #2196F3;
                color: white;
                min-width: 40px;
                min-height: 35px;
            }
            QPushButton:hover {
                background-color: #1976D2;
                border-color: #1976D2;
            }
            QPushButton:pressed {
                background-color: #0D47A1;
            }
        """)
        self.add_store_btn.clicked.connect(self.add_store_dialog)
        store_layout.addWidget(self.add_store_btn)
        
        # 修改店铺名称按钮
        self.edit_store_btn = QPushButton("✏️")
        self.edit_store_btn.setToolTip("修改当前店铺名称")
        self.edit_store_btn.setStyleSheet("""
            QPushButton {
                font-size: 16px;
                padding: 8px 12px;
                border: 2px solid #FF9800;
                border-radius: 6px;
                background-color: #FF9800;
                color: white;
                min-width: 40px;
                min-height: 35px;
            }
            QPushButton:hover {
                background-color: #F57C00;
                border-color: #F57C00;
            }
            QPushButton:pressed {
                background-color: #E65100;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                border-color: #999999;
                color: #666666;
            }
        """)
        self.edit_store_btn.clicked.connect(self.edit_store_dialog)
        self.edit_store_btn.setEnabled(False)  # 默认禁用，选择店铺后启用
        store_layout.addWidget(self.edit_store_btn)
        
        # 删除店铺按钮
        self.delete_store_btn = QPushButton("🗑️")
        self.delete_store_btn.setToolTip("删除当前店铺及其所有数据")
        self.delete_store_btn.setStyleSheet("""
            QPushButton {
                font-size: 16px;
                padding: 8px 12px;
                border: 2px solid #F44336;
                border-radius: 6px;
                background-color: #F44336;
                color: white;
                min-width: 40px;
                min-height: 35px;
            }
            QPushButton:hover {
                background-color: #D32F2F;
                border-color: #D32F2F;
            }
            QPushButton:pressed {
                background-color: #B71C1C;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                border-color: #999999;
                color: #666666;
            }
        """)
        self.delete_store_btn.clicked.connect(self.delete_store_dialog)
        self.delete_store_btn.setEnabled(False)  # 默认禁用，选择店铺后启用
        store_layout.addWidget(self.delete_store_btn)
        
        input_layout.addWidget(store_widget, 0, 1)

        order_label = QLabel("订单号：")
        order_label.setStyleSheet("margin-right: 2px; padding: 0px; font-size: 14px;")
        input_layout.addWidget(order_label, 0, 3)
        self.order_no_edit = QLineEdit()
        self.order_no_edit.setStyleSheet("margin-left: 0px; font-size: 14px;")
        self.order_no_edit.mousePressEvent = self.order_no_mouse_press
        input_layout.addWidget(self.order_no_edit, 0, 4, 1, 2)  # 订单号输入框跨2列，拉长宽度

        # 第一行第三个选项：退款金额
        refund_label = QLabel("退款金额：")
        refund_label.setStyleSheet("margin-right: 2px; padding: 0px; font-size: 14px;")
        input_layout.addWidget(refund_label, 0, 6)
        self.refund_amount_edit = QLineEdit()
        self.refund_amount_edit.setStyleSheet("margin-left: 0px; font-size: 14px;")
        self.refund_amount_edit.setPlaceholderText("¥")
        self.refund_amount_edit.setMaximumWidth(80)  # 退款金额输入框更窄
        self.refund_amount_edit.mousePressEvent = self.refund_amount_mouse_press
        input_layout.addWidget(self.refund_amount_edit, 0, 7)

        # 第二行：是否撤销、是否打款补偿、补偿金额、是否驳回、驳回结果、退款原因
        # 是否撤销和是否打款补偿上下排列 - 移到第二行开头
        checkbox_layout = QVBoxLayout()
        checkbox_layout.setSpacing(2)  # 减少复选框间距
        self.cancel_check = QCheckBox("是否撤销")
        self.cancel_check.setStyleSheet("font-size: 14px;")
        self.compensate_check = QCheckBox("是否打款补偿")
        self.compensate_check.setStyleSheet("font-size: 14px;")
        self.compensate_check.stateChanged.connect(self.toggle_comp_amount)
        checkbox_layout.addWidget(self.cancel_check)
        checkbox_layout.addWidget(self.compensate_check)
        input_layout.addLayout(checkbox_layout, 1, 0)  # 移到第0列

        comp_label = QLabel("补偿金额：")
        comp_label.setStyleSheet("margin-right: 2px; padding: 0px; font-size: 14px;")
        input_layout.addWidget(comp_label, 1, 1)
        self.comp_amount_edit = QLineEdit()
        self.comp_amount_edit.setStyleSheet("margin-left: 0px; font-size: 14px;")
        self.comp_amount_edit.setPlaceholderText("¥")
        self.comp_amount_edit.setMaximumWidth(80)  # 补偿金额输入框更窄
        self.comp_amount_edit.setEnabled(False)
        self.comp_amount_edit.mousePressEvent = self.comp_amount_mouse_press
        input_layout.addWidget(self.comp_amount_edit, 1, 2)  # 移到第2列

        # 是否驳回和驳回结果 - 移到补偿金额后面
        reject_layout = QVBoxLayout()
        reject_layout.setSpacing(2)  # 减少布局间距
        self.reject_check = QCheckBox("是否驳回")
        self.reject_check.setStyleSheet("font-size: 14px;")
        self.reject_check.stateChanged.connect(self.toggle_reject_result)
        reject_layout.addWidget(self.reject_check)
        
        self.reject_result_combo = QComboBox()
        self.reject_result_combo.setStyleSheet("font-size: 14px; margin-left: 0px;")
        self.reject_result_combo.addItems(["成功", "失败"])
        self.reject_result_combo.setEnabled(False)  # 默认不可选择
        reject_layout.addWidget(self.reject_result_combo)
        input_layout.addLayout(reject_layout, 1, 3)  # 移到第3列

        # 退款原因 - 移到和订单号一列（第3-4列）
        reason_label = QLabel("退款原因：")
        reason_label.setStyleSheet("margin-right: 2px; padding: 0px; font-size: 14px;")
        input_layout.addWidget(reason_label, 1, 4)
        self.reason_combo = QComboBox()
        self.reason_combo.setStyleSheet("margin-left: 0px; font-size: 14px;")
        reasons = ["商品腐败、变质、包装胀气等", "商品破损/压坏", "质量问题", "大小/规格/重量等与商品描述不符", "品种/标签/图片/包装等与商品描述不符", "货物与描述不符", "其他"]
        self.reason_combo.addItems(reasons)
        input_layout.addWidget(self.reason_combo, 1, 5, 1, 3)  # 跨3列显示，与订单号对齐

        # 第三行：备注输入框
        notes_label = QLabel("备注：")
        notes_label.setStyleSheet("margin-right: 2px; padding: 0px; font-size: 14px;")
        input_layout.addWidget(notes_label, 2, 0)
        self.notes_edit = QLineEdit()
        self.notes_edit.setStyleSheet("margin-left: 0px; font-size: 14px;")
        self.notes_edit.setPlaceholderText("请输入备注信息")
        input_layout.addWidget(self.notes_edit, 2, 1, 1, 6)  # 跨6列显示

        # 操作按钮 - 紧凑布局
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(2)  # 设置按钮间距为2像素，布局更紧凑
        
        self.add_btn = QPushButton("添加记录")
        self.add_btn.setStyleSheet("font-size: 14px; padding: 3px 8px;")
        self.update_btn = QPushButton("更新记录")
        self.update_btn.setStyleSheet("font-size: 14px; padding: 3px 8px;")
        self.delete_btn = QPushButton("删除选中")
        self.delete_btn.setStyleSheet("font-size: 14px; padding: 3px 8px;")
        self.clear_btn = QPushButton("清空输入")
        self.clear_btn.setStyleSheet("font-size: 14px; padding: 3px 8px;")
        self.import_btn = QPushButton("导入订单")
        self.import_btn.setStyleSheet("font-size: 14px; padding: 3px 8px;")
        self.export_btn = QPushButton("导出订单")
        self.export_btn.setStyleSheet("font-size: 14px; padding: 3px 8px;")
        
        btn_layout.addWidget(self.add_btn)
        btn_layout.addWidget(self.update_btn)
        btn_layout.addWidget(self.delete_btn)
        btn_layout.addWidget(self.clear_btn)
        btn_layout.addWidget(self.import_btn)
        btn_layout.addWidget(self.export_btn)
        btn_layout.addStretch()
        input_layout.addLayout(btn_layout, 3, 0, 1, 7)  # 修改为第3行

        # 绑定按钮事件
        self.add_btn.clicked.connect(self.add_record)
        self.update_btn.clicked.connect(self.update_record)
        self.delete_btn.clicked.connect(self.delete_record)
        self.clear_btn.clicked.connect(self.clear_input)
        self.import_btn.clicked.connect(self.import_excel)
        self.export_btn.clicked.connect(self.export_excel)

        # 刷新表格按钮 - 添加在添加记录按钮下方
        refresh_btn_layout = QHBoxLayout()
        refresh_btn_layout.setSpacing(2)  # 设置按钮间距为2像素，布局更紧凑
        
        self.refresh_table_btn = QPushButton("刷新表格")
        self.refresh_table_btn.setStyleSheet("font-size: 14px; padding: 3px 8px;")
        self.refresh_table_btn.clicked.connect(self.refresh_table_format)
        refresh_btn_layout.addWidget(self.refresh_table_btn)
        
        # 数据核对按钮
        self.check_data_btn = QPushButton("数据核对")
        self.check_data_btn.setFixedWidth(120)  # 设置固定宽度确保文字显示完整
        self.check_data_btn.setStyleSheet("""
            QPushButton {
                font-size: 14px; 
                padding: 3px 8px;
                background-color: #FF9800;
                color: white;
                border: 1px solid #F57C00;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
            QPushButton:pressed {
                background-color: #E65100;
            }
        """)
        self.check_data_btn.clicked.connect(self.check_data_consistency)
        refresh_btn_layout.addWidget(self.check_data_btn)
        
        refresh_btn_layout.addStretch()
        input_layout.addLayout(refresh_btn_layout, 4, 0, 1, 7)  # 添加在第4行

        # 顶部水平布局：信息录入区 + 店铺信息区
        top_horizontal_layout = QHBoxLayout()
        
        # 左侧：信息录入区
        top_horizontal_layout.addWidget(input_group)
        
        # 右侧：店铺信息区
        store_info_group = QGroupBox("店铺信息与统计")
        store_info_layout = QVBoxLayout()
        store_info_group.setLayout(store_info_layout)
        
        # 顶部：店铺名称显示、退款预算剩余和设置按钮
        top_row_layout = QHBoxLayout()
        
        # 店铺名称显示（直接显示信息录入板块选择的店铺）
        current_store_label_title = QLabel("当前店铺：")
        current_store_label_title.setStyleSheet("font-size: 16px; font-weight: bold;")
        top_row_layout.addWidget(current_store_label_title)
        self.current_store_label = QLabel("未选择")
        self.current_store_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #2E8B57;")
        top_row_layout.addWidget(self.current_store_label)
        
        # 添加弹性空间
        top_row_layout.addStretch()
        
        # 日退款预算剩余（移动到当前店铺和设置按钮之间）
        daily_budget_label_title = QLabel("退款预算剩余：")
        daily_budget_label_title.setStyleSheet("font-size: 14px; font-weight: bold;")
        top_row_layout.addWidget(daily_budget_label_title)
        self.daily_budget_remaining_label = QLabel("¥0.00")
        self.daily_budget_remaining_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #FF6B6B;")
        top_row_layout.addWidget(self.daily_budget_remaining_label)
        
        # 添加弹性空间
        top_row_layout.addStretch()
        
        # 店铺基本信息设置按钮（小齿轮图标）
        self.store_settings_btn = QPushButton("⚙")
        self.store_settings_btn.setStyleSheet("QPushButton { font-size: 18px; border: 1px solid #ccc; border-radius: 3px; }")
        self.store_settings_btn.setToolTip("店铺基本信息设置")
        self.store_settings_btn.clicked.connect(self.open_store_settings)
        top_row_layout.addWidget(self.store_settings_btn)
        
        store_info_layout.addLayout(top_row_layout)
        
        # 店铺统计信息（标题进一步增大）
        store_stats_label_title = QLabel("统计信息")
        store_stats_label_title.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50; margin-bottom: 8px;")
        store_info_layout.addWidget(store_stats_label_title)
        
        # 创建分割器容器，让用户可以自由拖拽调整大小
        splitter_container = QWidget()
        splitter_layout = QVBoxLayout(splitter_container)
        splitter_layout.setContentsMargins(2, 2, 2, 2)  # 减少边距，让内容更靠近边缘
        
        # 创建水平分割器（用于上下分割）
        horizontal_splitter = QSplitter(Qt.Horizontal)
        horizontal_splitter.setHandleWidth(6)  # 设置分割器手柄宽度
        horizontal_splitter.setStyleSheet("QSplitter::handle { background-color: #e0e0e0; }")
        
        # 创建左侧垂直分割器（用于左上和左下）
        left_vertical_splitter = QSplitter(Qt.Vertical)
        left_vertical_splitter.setHandleWidth(6)
        left_vertical_splitter.setStyleSheet("QSplitter::handle { background-color: #e0e0e0; }")
        
        # 创建右侧垂直分割器（用于右上和右下）
        right_vertical_splitter = QSplitter(Qt.Vertical)
        right_vertical_splitter.setHandleWidth(6)
        right_vertical_splitter.setStyleSheet("QSplitter::handle { background-color: #e0e0e0; }")
        
        # 第一组：退款数量统计（左上角，自由拉伸）
        refund_count_widget = QWidget()
        refund_count_layout = QVBoxLayout(refund_count_widget)
        refund_count_layout.setSpacing(4)
        refund_count_layout.setContentsMargins(8, 6, 8, 6)  # 增加内边距
        
        count_title = QLabel("退款数量统计")
        count_title.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50; margin: 0px;")
        refund_count_layout.addWidget(count_title)
        
        self.refund_count_label = QLabel("品质退款：0单\n其他退款：0单\n撤销品质退款：0单\n总退款率：0.00%")
        self.refund_count_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #343a40; margin: 0px; line-height: 18px;")
        self.refund_count_label.setWordWrap(True)
        self.refund_count_label.setMinimumHeight(90)  # 增加最小高度
        self.refund_count_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # 自由拉伸
        refund_count_layout.addWidget(self.refund_count_label)
        
        # 第二组：售后金额统计（右上角，自由拉伸）
        amount_widget = QWidget()
        amount_layout = QVBoxLayout(amount_widget)
        amount_layout.setSpacing(4)
        amount_layout.setContentsMargins(8, 6, 8, 6)  # 增加内边距
        
        amount_title = QLabel("售后金额统计")
        amount_title.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50; margin: 0px;")
        amount_layout.addWidget(amount_title)
        
        self.amount_label = QLabel("售后总金额：¥0.00\n售后金额占比：0.00%\n品质售后金额：¥0.00\n其他售后金额：¥0.00")
        self.amount_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #343a40; margin: 0px; line-height: 18px;")
        self.amount_label.setWordWrap(True)
        self.amount_label.setMinimumHeight(90)  # 增加最小高度
        self.amount_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # 自由拉伸
        amount_layout.addWidget(self.amount_label)
        
        # 第三组：品质退款率统计（左下角，自由拉伸）
        quality_widget = QWidget()
        quality_layout = QVBoxLayout(quality_widget)
        quality_layout.setSpacing(4)
        quality_layout.setContentsMargins(8, 6, 8, 6)  # 增加内边距
        
        quality_title = QLabel("品质退款率统计")
        quality_title.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50; margin: 0px;")
        quality_layout.addWidget(quality_title)
        
        self.quality_stats_label = QLabel("顾客申请品质退款率：0.00%\n实际计入品质退款率：0.00%\n品质退款撤销率：0.00%")
        self.quality_stats_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #343a40; margin: 0px; line-height: 18px;")
        self.quality_stats_label.setWordWrap(True)
        self.quality_stats_label.setMinimumHeight(90)  # 增加最小高度
        self.quality_stats_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # 自由拉伸
        quality_layout.addWidget(self.quality_stats_label)
        
        # 第四组：退款原因分析（右下角，自由拉伸）
        reason_widget = QWidget()
        reason_layout = QVBoxLayout(reason_widget)
        reason_layout.setSpacing(4)
        reason_layout.setContentsMargins(8, 6, 8, 6)  # 增加内边距
        
        reason_title = QLabel("退款原因分析")
        reason_title.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50; margin: 0px;")
        reason_layout.addWidget(reason_title)
        
        self.reason_analysis_label = QLabel("退款最多原因：无数据\n出现次数：0次\n占比：0.0%")
        self.reason_analysis_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #343a40; margin: 0px; line-height: 18px;")
        self.reason_analysis_label.setWordWrap(True)
        self.reason_analysis_label.setMinimumHeight(90)  # 增加最小高度
        self.reason_analysis_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # 自由拉伸
        reason_layout.addWidget(self.reason_analysis_label)
        
        # 将四个组件添加到分割器中
        left_vertical_splitter.addWidget(refund_count_widget)  # 左上
        left_vertical_splitter.addWidget(quality_widget)       # 左下
        right_vertical_splitter.addWidget(amount_widget)      # 右上
        right_vertical_splitter.addWidget(reason_widget)      # 右下
        
        # 设置左右分割器的初始比例
        left_vertical_splitter.setSizes([200, 200])
        right_vertical_splitter.setSizes([200, 200])
        
        # 将左右分割器添加到水平分割器
        horizontal_splitter.addWidget(left_vertical_splitter)
        horizontal_splitter.addWidget(right_vertical_splitter)
        
        # 设置水平分割器的初始比例
        horizontal_splitter.setSizes([400, 400])
        
        # 将水平分割器添加到容器布局
        splitter_layout.addWidget(horizontal_splitter)
        
        # 将分割器容器添加到主布局
        store_info_layout.addWidget(splitter_container)
        
        top_horizontal_layout.addWidget(store_info_group)
        
        # 左下角：搜索筛选区
        search_group = QGroupBox("搜索筛选")
        search_layout = QGridLayout()
        search_layout.setHorizontalSpacing(3)  # 设置水平间距为3像素，更紧凑
        search_layout.setVerticalSpacing(3)     # 设置垂直间距为3像素，更紧凑
        search_group.setLayout(search_layout)
        
        # 第一行：基础筛选条件
        search_layout.addWidget(QLabel("店铺："), 0, 0)
        self.search_store_combo = QComboBox()
        self.search_store_combo.addItem("全部")
        
        def on_store_changed(store_name):
            # 加载对应店铺的设置
            self.load_store_settings()
            # 触发搜索更新
            self.on_search_changed()
        
        self.search_store_combo.currentTextChanged.connect(on_store_changed)
        search_layout.addWidget(self.search_store_combo, 0, 1)

        search_layout.addWidget(QLabel("订单号："), 1, 0)
        self.search_order_edit = QLineEdit()
        self.search_order_edit.setMinimumWidth(100)  # 设置最小宽度，允许自动调整
        self.search_order_edit.textChanged.connect(self.on_search_changed)
        self.search_order_edit.mousePressEvent = self.search_order_mouse_press
        search_layout.addWidget(self.search_order_edit, 1, 1)

        # 退款原因多选下拉框
        search_layout.addWidget(QLabel("退款原因："), 2, 0)
        self.search_reason_combo = MultiSelectComboBox()
        reasons = ["商品腐败、变质、包装胀气等", "商品破损/压坏", "质量问题", "大小/规格/重量等与商品描述不符", "品种/标签/图片/包装等与商品描述不符", "货物与描述不符", "其他"]
        self.search_reason_combo.addItems(reasons)
        # 移除最大宽度限制，改为自动拉伸
        self.search_reason_combo.itemsChanged.connect(self.on_search_changed)
        search_layout.addWidget(self.search_reason_combo, 2, 1)

        # 登记日期 - 改为两行布局
        search_layout.addWidget(QLabel("开始日期："), 3, 0)
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDate(QDate.currentDate())
        self.start_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.start_date_edit.dateChanged.connect(self.on_search_changed)
        search_layout.addWidget(self.start_date_edit, 3, 1)
        
        search_layout.addWidget(QLabel("结束日期："), 4, 0)
        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDate(QDate.currentDate())
        self.end_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.end_date_edit.dateChanged.connect(self.on_search_changed)
        search_layout.addWidget(self.end_date_edit, 4, 1)

        # 其他筛选条件
        search_layout.addWidget(QLabel("撤销："), 5, 0)
        self.search_cancel_combo = QComboBox()
        self.search_cancel_combo.addItems(["全部", "是", "否"])
        self.search_cancel_combo.currentTextChanged.connect(self.on_search_changed)
        search_layout.addWidget(self.search_cancel_combo, 5, 1)

        search_layout.addWidget(QLabel("打款补偿："), 6, 0)
        self.search_compensate_combo = QComboBox()
        self.search_compensate_combo.addItems(["全部", "是", "否"])
        self.search_compensate_combo.currentTextChanged.connect(self.on_search_changed)
        search_layout.addWidget(self.search_compensate_combo, 6, 1)

        search_layout.addWidget(QLabel("驳回："), 7, 0)
        self.search_reject_combo = QComboBox()
        self.search_reject_combo.addItems(["全部", "是", "否"])
        self.search_reject_combo.currentTextChanged.connect(self.on_search_changed)
        search_layout.addWidget(self.search_reject_combo, 7, 1)

        search_layout.addWidget(QLabel("驳回结果："), 8, 0)
        self.search_reject_result_combo = QComboBox()
        self.search_reject_result_combo.addItems(["全部", "成功", "失败"])
        self.search_reject_result_combo.currentTextChanged.connect(self.on_search_changed)
        search_layout.addWidget(self.search_reject_result_combo, 8, 1)

        # 重置和显示全部按钮（一行两个按钮）
        button_layout = QHBoxLayout()
        
        reset_btn = QPushButton("重置筛选")
        reset_btn.setFixedWidth(100)  # 设置固定宽度
        reset_btn.clicked.connect(self.reset_search)
        button_layout.addWidget(reset_btn)
        
        show_all_btn = QPushButton("显示全部")
        show_all_btn.setFixedWidth(100)  # 设置固定宽度
        show_all_btn.clicked.connect(self.show_all_records)
        button_layout.addWidget(show_all_btn)
        
        search_layout.addLayout(button_layout, 9, 0, 1, 2)
        
        # 快捷日期按钮 - 改为每个按钮一行
        quick_date_group = QGroupBox("快捷日期")
        quick_date_layout = QVBoxLayout()
        quick_date_layout.setSpacing(5)  # 按钮间距5像素
        quick_date_group.setLayout(quick_date_layout)
        
        # 每个按钮独占一行
        today_btn = QPushButton("今天")
        today_btn.clicked.connect(lambda: self.set_quick_date(0))
        quick_date_layout.addWidget(today_btn)
        
        yesterday_btn = QPushButton("昨天")
        yesterday_btn.clicked.connect(lambda: self.set_quick_date(1))
        quick_date_layout.addWidget(yesterday_btn)
        
        # 前一天和后一天按钮
        prev_next_layout = QHBoxLayout()
        prev_next_layout.setSpacing(2)
        
        prev_day_btn = QPushButton("◀ 前一天")
        prev_day_btn.setStyleSheet("font-size: 12px; padding: 2px 6px;")
        prev_day_btn.clicked.connect(self.previous_day)
        prev_next_layout.addWidget(prev_day_btn)
        
        next_day_btn = QPushButton("后一天 ▶")
        next_day_btn.setStyleSheet("font-size: 12px; padding: 2px 6px;")
        next_day_btn.clicked.connect(self.next_day)
        prev_next_layout.addWidget(next_day_btn)
        
        quick_date_layout.addLayout(prev_next_layout)
        
        week_btn = QPushButton("近7天")
        week_btn.clicked.connect(lambda: self.set_quick_date(7))
        quick_date_layout.addWidget(week_btn)
        
        month_btn = QPushButton("近30天")
        month_btn.clicked.connect(lambda: self.set_quick_date(30))
        quick_date_layout.addWidget(month_btn)
        
        # 右下角：订单记录表格
        table_group = QGroupBox("订单记录表格")
        table_layout = QVBoxLayout(table_group)
        
        # 调试信息标签
        self.debug_label = QLabel("表格区域 - 显示筛选后的订单记录")
        self.debug_label.setStyleSheet("color: #666; font-size: 10px; padding: 5px; background-color: #f0f0f0; border: 1px solid #ccc;")
        table_layout.addWidget(self.debug_label)
        
        self.table = QTableWidget()
        self.table.setColumnCount(11)  # 恢复为11列
        self.table.setHorizontalHeaderLabels(["店铺名称", "订单号", "退款原因", "退款金额", "撤销", "打款补偿", "补偿金额", "驳回", "驳回结果", "登记日期", "备注"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)  # 设置扩展选择模式，支持多选和Ctrl+A
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)  # 禁用编辑，使用双击切换功能
        
        # 设置列宽自适应模式
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)  # 设置为交互模式
        
        # 为订单号、退款原因列设置特殊拉伸模式，确保字符显示完整
        self.table.setColumnWidth(1, 200)  # 订单号列设置较宽宽度
        self.table.setColumnWidth(2, 250)  # 退款原因列设置较宽宽度
        
        # 其他列使用默认宽度
        self.table.setColumnWidth(0, 120)  # 店铺名称
        self.table.setColumnWidth(3, 100)  # 退款金额
        self.table.setColumnWidth(4, 60)   # 撤销
        self.table.setColumnWidth(5, 80)   # 打款补偿
        self.table.setColumnWidth(6, 100)  # 补偿金额
        self.table.setColumnWidth(7, 60)   # 驳回
        self.table.setColumnWidth(8, 100)  # 驳回结果
        self.table.setColumnWidth(9, 100)   # 登记日期（现在在第9列）
        
        # 设置列宽调整策略
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # 订单号：根据内容调整
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # 退款原因：根据内容调整
        header.setSectionResizeMode(10, QHeaderView.Stretch)  # 备注：完全自动拉伸（现在在第10列）
        
        # 设置自定义的编辑检查函数
        self.table.setItemDelegate(CustomItemDelegate(self))
        self.table.itemDoubleClicked.connect(self.on_item_double_clicked)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        # 连接单元格编辑完成信号
        self.table.cellChanged.connect(self.on_cell_changed)
        table_layout.addWidget(self.table)
        
        # 将区域添加到分割器中
        # 上部区域：信息录入区（左）和店铺信息区（右）
        top_splitter.addWidget(input_group)           # 左：信息录入区
        top_splitter.addWidget(store_info_group)      # 右：店铺信息区
        
        # 删除上部区域最小宽度限制，让用户完全自由调整
        
        # 下部左侧：垂直布局（搜索筛选区 + 快捷日期）
        left_bottom_widget = QWidget()
        left_bottom_layout = QVBoxLayout(left_bottom_widget)
        left_bottom_layout.addWidget(search_group)
        left_bottom_layout.addWidget(quick_date_group)
        
        # 下部区域：左侧（搜索筛选）和右侧（订单记录表格）
        bottom_splitter.addWidget(left_bottom_widget)  # 左：搜索筛选区
        bottom_splitter.addWidget(table_group)         # 右：订单记录表格
        
        # 删除下部区域最小宽度限制，让用户完全自由调整

        # 底部状态栏
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

        # 加载保存的界面设置
        self.load_window_settings()

        # 初始化店铺信息下拉框
        self.load_store_info_combo()

    def on_store_combo_changed(self, store_name):
        """信息录入区店铺选择变化"""
        # 不再同步到搜索筛选区，保持两个区域独立
        
        # 启用/禁用编辑和删除按钮
        if store_name and store_name != "请先添加店铺":
            # 启用编辑和删除按钮（只有在选择真实店铺时）
            if hasattr(self, 'edit_store_btn') and hasattr(self, 'delete_store_btn'):
                self.edit_store_btn.setEnabled(True)
                self.delete_store_btn.setEnabled(True)
        else:
            # 禁用编辑和删除按钮
            if hasattr(self, 'edit_store_btn') and hasattr(self, 'delete_store_btn'):
                self.edit_store_btn.setEnabled(False)
                self.delete_store_btn.setEnabled(False)

    def sync_store_selection(self, store_name):
        """同步所有店铺选择框"""
        # 同步搜索区的店铺选择
        if hasattr(self, 'search_store_combo'):
            index = self.search_store_combo.findText(store_name)
            if index >= 0:
                self.search_store_combo.setCurrentIndex(index)
        
        # 同步店铺信息区的店铺显示（现在直接显示搜索筛选区选择的店铺）
        # 加载对应店铺的设置
        self.load_store_settings()
        self.update_store_stats_display()

    def save_window_settings(self):
        """空方法，已删除记忆功能"""
        # 不再保存窗口设置，使用固定的默认值
        pass

    def load_window_settings(self):
        """设置固定的默认窗口设置（删除记忆功能）"""
        # 直接设置固定的默认值，不使用记忆功能
        self.main_splitter.setSizes([200, 750])  # 主分割器固定比例：上200，下750
        self.top_splitter.setSizes([1000, 400])  # 上部水平分割器固定比例：左1000，右400
        self.bottom_splitter.setSizes([100, 1300]) # 下部水平分割器固定比例：左100，右1300

    def closeEvent(self, event):
        """窗口关闭事件，保存设置"""
        self.save_window_settings()
        event.accept()

    def load_store_settings(self):
        """从数据库加载店铺设置"""
        # 获取当前选择的店铺
        current_store = None
        if hasattr(self, 'search_store_combo') and self.search_store_combo.currentText():
            current_store = self.search_store_combo.currentText()
        
        # 只在程序启动时设置默认店铺（第一次调用时）
        if not hasattr(self, '_store_settings_initialized'):
            self._store_settings_initialized = True
            if not current_store:
                # 默认选择"全部"店铺
                current_store = "全部"
                # 如果搜索筛选区已初始化，更新选择
                if hasattr(self, 'search_store_combo'):
                    index = self.search_store_combo.findText("全部")
                    if index >= 0:
                        self.search_store_combo.setCurrentIndex(index)
        
        if current_store and current_store != "全部":
            # 获取店铺ID
            stores = self.db.get_stores()
            store_id = None
            for sid, sname in stores:
                if sname == current_store:
                    store_id = sid
                    break
            
            if store_id:
                # 从数据库加载设置
                db_settings = self.db.get_store_settings(store_id)
                if db_settings:
                    self.store_settings = db_settings
                else:
                    # 如果没有设置，使用默认值
                    self.store_settings = {'daily_orders': 0, 'daily_sales': 0.0, 'refund_budget': 0.0}
        else:
            # 如果选择了"全部"店铺，从全局设置加载
            self.store_settings = self.db.get_global_settings()

    def load_store_info_combo(self):
        """加载店铺信息下拉框（现在使用信息录入区的店铺选择）"""
        # 不再需要这个功能，因为店铺信息区直接显示信息录入区选择的店铺
        pass

    def on_store_info_changed(self, store_name):
        """店铺信息选择变化"""
        if store_name:
            # 获取预估订单量
            estimated_orders = self.db.get_estimated_orders(store_name)
            self.estimated_orders_edit.setText(str(estimated_orders) if estimated_orders > 0 else "")
            
            # 更新店铺统计
            self.update_store_stats(store_name)

    def update_store_stats(self, store_name):
        """更新店铺统计信息"""
        if not store_name:
            return
            
        # 获取店铺退款统计（排除撤销订单）
        stats = self.db.get_store_refund_stats(store_name)
        refund_count = stats['refund_count']
        total_refund = stats['total_refund']
        total_comp = stats['total_comp']
        
        # 获取预估订单量（实时更新到数据库）
        estimated_text = self.estimated_orders_edit.text().strip()
        estimated_orders = 0
        if estimated_text:
            try:
                estimated_orders = int(estimated_text)
                if estimated_orders > 0:
                    # 实时保存到数据库
                    self.db.set_estimated_orders(store_name, estimated_orders)
            except ValueError:
                estimated_orders = self.db.get_estimated_orders(store_name)
        else:
            estimated_orders = self.db.get_estimated_orders(store_name)
        
        # 计算退款率
        refund_rate = 0.0
        if estimated_orders > 0:
            refund_rate = (refund_count / estimated_orders) * 100
        
        # 更新显示
        self.store_stats_label.setText(
            f"有效退款：{refund_count}单 | 退款率：{refund_rate:.2f}% | 总金额：¥{total_refund + total_comp:.2f}"
        )

    def update_refund_rate(self):
        """更新退款率显示（现在使用信息录入区的店铺选择）"""
        store_name = self.store_combo.currentText()
        if store_name:
            self.update_store_stats(store_name)

    def open_store_settings(self):
        """打开店铺基本信息设置对话框"""
        # 确保当前店铺设置已从数据库加载
        self.load_store_settings()
        dialog = StoreSettingsDialog(self)
        dialog.load_settings(self.store_settings)
        if dialog.exec_() == QDialog.Accepted:
            # 设置已保存，更新显示
            self.update_store_stats_display()

    def update_store_stats_display(self):
        """更新店铺统计信息显示"""
        # 更新当前店铺名称显示（使用搜索筛选区的店铺选择）
        current_store = self.search_store_combo.currentText() if self.search_store_combo.currentText() else "未选择"
        self.current_store_label.setText(current_store)
        
        # 更新日退款预算剩余
        daily_budget_remaining = self.calculate_daily_budget_remaining()
        self.daily_budget_remaining_label.setText(f"¥{daily_budget_remaining:.2f}")
        
        # 更新增强的退款统计信息
        enhanced_stats = self.calculate_enhanced_refund_stats()
        
        # 更新品质退款相关统计
        quality_stats = self.calculate_quality_refund_stats()
        
        # 更新售后金额相关统计
        refund_stats = self.calculate_refund_amount_stats()
        
        # 第一组：退款数量统计（完整显示）
        self.refund_count_label.setText(
            f"品质退款：{enhanced_stats['quality_refund_count']}单\n"
            f"其他退款：{enhanced_stats['other_refund_count']}单\n"
            f"撤销品质退款：{enhanced_stats['canceled_quality_count']}单\n"
            f"总退款率：{enhanced_stats['total_refund_rate']:.2f}%"
        )
        
        # 第二组：金额统计（完整显示）
        self.amount_label.setText(
            f"售后总金额：¥{refund_stats['total_refund']:.2f}\n"
            f"售后金额占比：{refund_stats['refund_ratio']:.2f}%\n"
            f"品质售后金额：¥{enhanced_stats['quality_after_sales_amount']:.2f}\n"
            f"其他售后金额：¥{enhanced_stats['other_after_sales_amount']:.2f}"
        )
        
        # 第三组：品质退款率统计（完整显示）
        self.quality_stats_label.setText(
            f"顾客申请品质退款率：{quality_stats['apply_rate']:.2f}%\n"
            f"实际计入品质退款率：{quality_stats['actual_rate']:.2f}%\n"
            f"品质退款撤销率：{quality_stats['cancel_rate']:.2f}%"
        )
        
        # 第四组：退款原因分析（完整显示）
        self.reason_analysis_label.setText(
            f"退款最多原因：{enhanced_stats['top_refund_reason']}\n"
            f"出现次数：{enhanced_stats['top_reason_count']}次\n"
            f"占比：{enhanced_stats['top_reason_ratio']:.1f}%"
        )

    def calculate_daily_budget_remaining(self):
        """计算日退款预算剩余（支持多天筛选）"""
        if not self.store_settings:
            return 0.0
        
        # 获取筛选的天数
        start_date = self.start_date_edit.date().toPyDate()
        end_date = self.end_date_edit.date().toPyDate()
        days_count = (end_date - start_date).days + 1  # 计算筛选的天数
        
        # 获取用户设置的日退款预算
        daily_refund_budget = self.store_settings.get('refund_budget', 0.0)
        
        # 计算多天的总预算
        total_refund_budget = daily_refund_budget * days_count
        
        # 计算筛选期间的总退款金额
        filtered_records = self.get_filtered_records()
        total_refund = 0.0
        for record in filtered_records:
            # 打款金额始终计入（不管撤销驳回与否）
            if record['compensate']:  # 已打款补偿
                total_refund += record['comp_amount']
            
            # 退款金额计算：只计算未撤销且未驳回成功的订单的退款金额
            if not record['cancel'] and not (record.get('reject') and record.get('reject_result') == "成功"):  # 未撤销且未驳回成功
                total_refund += record['refund_amount']
        
        return max(0.0, total_refund_budget - total_refund)

    def calculate_today_refund_amount(self, date):
        """计算指定日期的退款金额（基于当前筛选条件）"""
        # 使用与表格相同的筛选条件获取记录
        filtered_records = self.get_filtered_records()
        
        # 筛选指定日期的记录
        today_str = date.strftime('%Y-%m-%d')
        today_records = [r for r in filtered_records if r['record_date'] == today_str]
        
        if not today_records:
            return 0.0
        
        # 计算今天的退款金额（使用与退款金额统计相同的逻辑）
        today_refund = 0.0
        for record in today_records:
            # 打款金额始终计入（不管撤销驳回与否）
            if record['compensate']:  # 已打款补偿
                today_refund += record['comp_amount']
            
            # 退款金额计算：只计算未撤销且未驳回成功的订单的退款金额
            if not record['cancel'] and not (record.get('reject') and record.get('reject_result') == "成功"):  # 未撤销且未驳回成功
                today_refund += record['refund_amount']
        
        return today_refund

    def calculate_quality_refund_stats(self):
        """计算品质退款相关统计（基于当前筛选条件，支持多天筛选）"""
        # 品质退款原因列表
        quality_reasons = [
            "商品腐败、变质、包装胀气等", 
            "商品破损/压坏", 
            "质量问题",
            "大小/规格/重量等与商品描述不符",
            "品种/标签/图片/包装等与商品描述不符",
            "货物与描述不符"
        ]
        
        # 使用与表格相同的筛选条件获取记录
        filtered_records = self.get_filtered_records()
        
        if not filtered_records:
            return {'apply_rate': 0.0, 'actual_rate': 0.0, 'cancel_rate': 0.0}
        
        # 获取筛选的天数
        start_date = self.start_date_edit.date().toPyDate()
        end_date = self.end_date_edit.date().toPyDate()
        days_count = (end_date - start_date).days + 1  # 计算筛选的天数
        
        # 计算品质退款相关统计
        # 顾客申请品质退款率：品质退款订单数 ÷ (用户设置的日订单量 × 筛选天数)
        daily_orders = self.store_settings.get('daily_orders', 0)
        total_orders = daily_orders * days_count  # 多天的总订单量
        
        # 顾客申请品质退款率：只要退款原因不是"其他"的都算
        quality_refund_orders = [r for r in filtered_records if r['reason'] in quality_reasons]
        apply_quality_count = len(quality_refund_orders)
        apply_rate = (apply_quality_count / total_orders * 100) if total_orders > 0 else 0.0
        
        # 实际计入品质退款率：减去已撤销和驳回成功的订单
        actual_quality_count = apply_quality_count
        for record in quality_refund_orders:
            if record['cancel']:  # 已撤销
                actual_quality_count -= 1
            elif record.get('reject') and record.get('reject_result') == "成功":  # 驳回成功
                actual_quality_count -= 1
        
        # 修复：实际计入品质退款率应该使用与顾客申请相同的分母（total_orders）
        actual_rate = (actual_quality_count / total_orders * 100) if total_orders > 0 else 0.0
        
        # 品质退款撤销率：已撤销的品质退款订单数 ÷ 总品质退款订单数
        canceled_quality_count = sum(1 for r in quality_refund_orders if r['cancel'])
        cancel_rate = (canceled_quality_count / apply_quality_count * 100) if apply_quality_count > 0 else 0.0
        
        return {
            'apply_rate': apply_rate,    # 顾客申请品质退款率
            'actual_rate': actual_rate,  # 实际计入品质退款率
            'cancel_rate': cancel_rate   # 品质退款撤销率
        }

    def calculate_refund_amount_stats(self):
        """计算退款金额相关统计（基于当前筛选条件，支持多天筛选）"""
        # 使用与表格相同的筛选条件获取记录
        filtered_records = self.get_filtered_records()
        
        if not filtered_records:
            return {'total_refund': 0.0, 'refund_ratio': 0.0}
        
        # 获取筛选的天数
        start_date = self.start_date_edit.date().toPyDate()
        end_date = self.end_date_edit.date().toPyDate()
        days_count = (end_date - start_date).days + 1  # 计算筛选的天数
        
        # 计算退款金额
        total_refund = 0.0
        for record in filtered_records:
            # 打款金额始终计入（不管撤销驳回与否）
            if record['compensate']:  # 已打款补偿
                total_refund += record['comp_amount']
            
            # 退款金额计算：只计算未撤销且未驳回成功的订单的退款金额
            if not record['cancel'] and not (record.get('reject') and record.get('reject_result') == "成功"):  # 未撤销且未驳回成功
                total_refund += record['refund_amount']
        
        # 计算退款金额占比：退款金额 ÷ (用户设置的日销售额 × 筛选天数)
        daily_sales = self.store_settings.get('daily_sales', 0.0)
        total_sales = daily_sales * days_count  # 多天的总销售额
        refund_ratio = (total_refund / total_sales * 100) if total_sales > 0 else 0.0
        
        return {
            'total_refund': total_refund,  # 退款金额
            'refund_ratio': refund_ratio  # 退款金额占比
        }

    def calculate_enhanced_refund_stats(self):
        """计算增强的退款统计信息（基于当前筛选条件）"""
        # 使用与表格相同的筛选条件获取记录
        filtered_records = self.get_filtered_records()
        
        if not filtered_records:
            return {
                'quality_refund_count': 0,
                'other_refund_count': 0,
                'canceled_quality_count': 0,
                'total_refund_rate': 0.0,
                'quality_after_sales_amount': 0.0,
                'other_after_sales_amount': 0.0,
                'top_refund_reason': '无数据',
                'top_reason_count': 0,
                'top_reason_ratio': 0.0
            }
        
        # 品质退款原因列表
        quality_reasons = [
            "商品腐败、变质、包装胀气等", 
            "商品破损/压坏", 
            "质量问题",
            "大小/规格/重量等与商品描述不符",
            "品种/标签/图片/包装等与商品描述不符",
            "货物与描述不符"
        ]
        
        # 获取筛选的天数
        start_date = self.start_date_edit.date().toPyDate()
        end_date = self.end_date_edit.date().toPyDate()
        days_count = (end_date - start_date).days + 1  # 计算筛选的天数
        
        # 计算总订单数
        daily_orders = self.store_settings.get('daily_orders', 0)
        total_orders = daily_orders * days_count  # 多天的总订单量
        
        # 统计退款原因出现次数
        reason_counts = {}
        
        # 初始化统计变量
        quality_refund_count = 0
        other_refund_count = 0
        canceled_quality_count = 0
        quality_after_sales_amount = 0.0
        other_after_sales_amount = 0.0
        
        # 遍历所有记录进行统计
        for record in filtered_records:
            reason = record['reason']
            
            # 统计退款原因出现次数
            reason_counts[reason] = reason_counts.get(reason, 0) + 1
            
            # 判断是否为品质退款
            is_quality_refund = reason in quality_reasons
            
            # 统计数量
            if is_quality_refund:
                quality_refund_count += 1
                # 统计撤销的品质退款数量
                if record['cancel']:
                    canceled_quality_count += 1
            else:
                other_refund_count += 1
            
            # 计算售后金额（只计算未撤销且未驳回成功的订单）
            if not record['cancel'] and not (record.get('reject') and record.get('reject_result') == "成功"):
                # 计算该订单的售后金额（退款金额 + 打款补偿金额）
                order_after_sales_amount = record['refund_amount']
                if record['compensate']:
                    order_after_sales_amount += record['comp_amount']
                
                # 根据退款类型累加到对应的售后金额
                if is_quality_refund:
                    quality_after_sales_amount += order_after_sales_amount
                else:
                    other_after_sales_amount += order_after_sales_amount
        
        # 计算总退款率
        total_refund_count = quality_refund_count + other_refund_count
        total_refund_rate = (total_refund_count / total_orders * 100) if total_orders > 0 else 0.0
        
        # 找出退款最多的原因
        top_refund_reason = '无数据'
        top_reason_count = 0
        top_reason_ratio = 0.0
        
        if reason_counts:
            top_refund_reason = max(reason_counts, key=reason_counts.get)
            top_reason_count = reason_counts[top_refund_reason]
            top_reason_ratio = (top_reason_count / total_refund_count * 100) if total_refund_count > 0 else 0.0
        
        return {
            'quality_refund_count': quality_refund_count,        # 品质退款数量
            'other_refund_count': other_refund_count,            # 其他退款数量
            'canceled_quality_count': canceled_quality_count,    # 撤销的品质退款数量
            'total_refund_rate': total_refund_rate,              # 总退款率
            'quality_after_sales_amount': quality_after_sales_amount,  # 品质售后金额
            'other_after_sales_amount': other_after_sales_amount,      # 其他售后金额
            'top_refund_reason': top_refund_reason,              # 退款最多的原因
            'top_reason_count': top_reason_count,                # 最多原因出现次数
            'top_reason_ratio': top_reason_ratio                 # 最多原因占比
        }

    def update_total_amount_display(self):
        """更新右上角全局统计显示"""
        # 获取所有记录
        records = self.db.get_all_records()
        
        # 计算总金额（排除撤销订单）
        total_refund = sum(r['refund_amount'] for r in records if not r['cancel'])
        total_comp = sum(r['comp_amount'] for r in records if not r['cancel'])
        total_amount = total_refund + total_comp
        
        # 计算总退款率
        total_refund_count = sum(1 for r in records if not r['cancel'])
        total_estimated_orders = self.get_total_estimated_orders()
        
        total_refund_rate = 0.0
        if total_estimated_orders > 0:
            total_refund_rate = (total_refund_count / total_estimated_orders) * 100
        
        # 更新右上角全局统计显示
        if hasattr(self, 'global_stats_label'):
            self.global_stats_label.setText(f"总金额：¥{total_amount:,.2f} | 总退款率：{total_refund_rate:.2f}%")

    def get_total_estimated_orders(self):
        """获取所有店铺的预估订单量总和"""
        stores = self.db.get_stores()
        total_estimated = 0
        for store_id, store_name in stores:
            estimated = self.db.get_estimated_orders(store_name)
            total_estimated += estimated
        return total_estimated

    def apply_stylesheet(self):
        """应用极简风格样式表"""
        stylesheet = """
        /* 主窗口背景色 */
        QMainWindow {
            background-color: #F8F9FA;  /* 极浅灰色背景 */
        }
        
        /* 中央控件背景 */
        QWidget {
            background-color: #F8F9FA;
            color: #212529;  /* 深灰文字 */
        }
        
        /* 分组框样式 */
        QGroupBox {
            font-weight: bold;
            font-size: 12px;
            border: 1px solid #DEE2E6;  /* 浅灰边框 */
            border-radius: 6px;
            margin-top: 8px;
            padding-top: 8px;
            background-color: #FFFFFF;  /* 白色背景 */
        }
        
        QGroupBox::title {
            subcontrol-origin: margin;
            subcontrol-position: top center;
            padding: 0 6px;
            background-color: #6C757D;  /* 中灰标题背景 */
            color: white;
            border-radius: 3px;
        }
        
        /* 按钮样式 */
        QPushButton {
            background-color: #6C757D;  /* 中灰背景 */
            color: white;
            border: none;
            padding: 6px 12px;
            border-radius: 3px;
            font-weight: normal;
            min-width: 80px;
        }
        
        QPushButton:hover {
            background-color: #5A6268;  /* 深灰悬停 */
        }
        
        QPushButton:pressed {
            background-color: #495057;  /* 更深灰按下 */
        }
        
        /* 重要操作按钮特殊样式 */
        QPushButton[important="true"] {
            background-color: #DC3545;  /* 红色强调 */
        }
        
        QPushButton[important="true"]:hover {
            background-color: #C82333;
        }
        
        /* 输入框样式 */
        QLineEdit, QComboBox, QDateEdit {
            padding: 6px;
            border: 1px solid #CED4DA;
            border-radius: 4px;
            background-color: white;
            selection-background-color: #6C757D;
        }
        
        QLineEdit:focus, QComboBox:focus, QDateEdit:focus {
            border: 2px solid #6C757D;
        }
        
        /* 表格样式 */
        QTableWidget {
            gridline-color: #E9ECEF;
            border: 1px solid #CED4DA;
            border-radius: 4px;
            background-color: white;
        }
        
        QTableWidget::item {
            padding: 6px;
            border-bottom: 1px solid #F8F9FA;
        }
        
        QTableWidget::item:selected {
            background-color: #6C757D;  /* 中灰选中 */
            color: white;
        }
        
        QHeaderView::section {
            background-color: #6C757D;  /* 中灰表头 */
            color: white;
            padding: 8px;
            border: none;
            font-weight: bold;
        }
        
        /* 复选框样式 */
        QCheckBox {
            spacing: 8px;
        }
        
        QCheckBox::indicator {
            width: 16px;
            height: 16px;
        }
        
        QCheckBox::indicator:unchecked {
            border: 1px solid #CCCCCC;
            background-color: white;
            border-radius: 2px;
        }
        
        QCheckBox::indicator:checked {
            border: 1px solid #2E8B57;
            background-color: #2E8B57;
            border-radius: 2px;
        }
        
        /* 状态栏样式 */
        QStatusBar {
            background-color: #2E8B57;
            color: white;
            padding: 4px;
        }
        
        /* 标签样式 */
        QLabel {
            color: #333333;
            font-weight: normal;
        }
        """
        self.setStyleSheet(stylesheet)

    def setup_shortcuts(self):
        QShortcut(QKeySequence("Ctrl+S"), self, self.add_record)
        QShortcut(QKeySequence("Ctrl+E"), self, self.export_excel)
        QShortcut(QKeySequence("Ctrl+F"), self, lambda: self.search_order_edit.setFocus())
        QShortcut(QKeySequence("Ctrl+D"), self, self.delete_record)

    def toggle_comp_amount(self, state):
        self.comp_amount_edit.setEnabled(state == Qt.Checked)
        if state != Qt.Checked:
            self.comp_amount_edit.clear()

    def toggle_reject_result(self, state):
        """控制驳回结果下拉框的可用性"""
        self.reject_result_combo.setEnabled(state == Qt.Checked)
        if state != Qt.Checked:
            self.reject_result_combo.setCurrentIndex(0)  # 重置为默认值

    def load_stores(self):
        """加载店铺列表到所有下拉框"""
        # 清空所有店铺下拉框
        self.store_combo.clear()
        self.search_store_combo.clear()
            
        stores = self.db.get_stores()
        self.search_store_combo.addItem("全部")
        for store_id, store_name in stores:
            self.store_combo.addItem(store_name, store_id)
            self.search_store_combo.addItem(store_name, store_id)
        
        # 如果有店铺，信息录入区选择第一个，搜索筛选区选择"全部"
        if self.store_combo.count() > 0:
            # 信息录入区选择第一个店铺
            current_store = self.store_combo.currentText()
            # 搜索筛选区选择"全部"
            self.search_store_combo.setCurrentIndex(0)  # 0是"全部"选项
            # 不再同步店铺信息显示，保持两个区域独立
        
        if self.store_combo.count() == 0:
            self.store_combo.addItem("请先添加店铺", None)

    def add_store_dialog(self):
        """添加店铺对话框"""
        name, ok = QInputDialog.getText(self, "添加店铺", "店铺名称：")
        if ok and name.strip():
            name = name.strip()
            if self.db.add_store(name):
                self.load_stores()
                self.show_tooltip(f"店铺 {name} 已添加", "rgba(76, 175, 80, 0.95)", 1500)  # 绿色气泡显示1.5秒
            else:
                QMessageBox.warning(self, "错误", f"店铺 '{name}' 已存在！")

    def edit_store_dialog(self):
        """修改店铺名称对话框"""
        current_store = self.store_combo.currentText()
        if not current_store or current_store == "请先添加店铺":
            QMessageBox.warning(self, "错误", "请先选择一个店铺！")
            return
        
        # 获取当前店铺ID
        stores = self.db.get_stores()
        store_id = None
        for sid, sname in stores:
            if sname == current_store:
                store_id = sid
                break
        
        if not store_id:
            QMessageBox.warning(self, "错误", "未找到选中的店铺！")
            return
        
        new_name, ok = QInputDialog.getText(self, "修改店铺名称", "新店铺名称：", text=current_store)
        if ok and new_name.strip():
            new_name = new_name.strip()
            if new_name == current_store:
                QMessageBox.information(self, "提示", "店铺名称未改变！")
                return
            
            if self.db.update_store_name(store_id, new_name):
                self.load_stores()
                # 更新当前选择
                index = self.store_combo.findText(new_name)
                if index >= 0:
                    self.store_combo.setCurrentIndex(index)
                self.show_tooltip(f"店铺名称已修改为 {new_name}", "rgba(33, 150, 243, 0.95)", 1500)  # 蓝色气泡
            else:
                QMessageBox.warning(self, "错误", f"店铺名称 '{new_name}' 已存在或修改失败！")

    def delete_store_dialog(self):
        """删除店铺对话框"""
        current_store = self.store_combo.currentText()
        if not current_store or current_store == "请先添加店铺":
            QMessageBox.warning(self, "错误", "请先选择一个店铺！")
            return
        
        # 获取当前店铺ID
        stores = self.db.get_stores()
        store_id = None
        for sid, sname in stores:
            if sname == current_store:
                store_id = sid
                break
        
        if not store_id:
            QMessageBox.warning(self, "错误", "未找到选中的店铺！")
            return
        
        # 确认删除对话框
        reply = QMessageBox.question(self, "确认删除", 
                                    f"确定要删除店铺 '{current_store}' 吗？\n\n⚠️ 警告：删除后该店铺的所有退款记录也将被删除！\n此操作不可撤销！",
                                    QMessageBox.Yes | QMessageBox.No, 
                                    QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            if self.db.delete_store(store_id):
                self.load_stores()
                # 重置选择
                if self.store_combo.count() > 0:
                    self.store_combo.setCurrentIndex(0)
                else:
                    self.store_combo.addItem("请先添加店铺", None)
                    self.store_combo.setCurrentIndex(0)
                
                # 禁用编辑和删除按钮
                if hasattr(self, 'edit_store_btn') and hasattr(self, 'delete_store_btn'):
                    self.edit_store_btn.setEnabled(False)
                    self.delete_store_btn.setEnabled(False)
                
                self.show_tooltip(f"店铺 {current_store} 及其所有数据已删除", "rgba(244, 67, 54, 0.95)", 2000)  # 红色气泡
            else:
                QMessageBox.warning(self, "错误", "删除店铺失败！")

    def get_current_date(self):
        return datetime.now().strftime("%Y-%m-%d")

    def update_debug_label(self, record_count, order_no, reason, store_name):
        """更新调试标签显示当前筛选结果"""
        debug_text = f"表格区域 - 当前显示 {record_count} 条订单记录"
        
        # 如果有筛选条件，显示筛选信息
        conditions = []
        if order_no:
            conditions.append(f"订单号: {order_no}")
        if reason and reason != "全部":
            conditions.append(f"退款原因: {reason}")
        if store_name and store_name != "全部":
            conditions.append(f"店铺: {store_name}")
        
        if conditions:
            debug_text += f" | 筛选条件: {' | '.join(conditions)}"
        
        self.debug_label.setText(debug_text)

    def parse_date_string(self, date_str):
        """解析多种日期格式，支持带时间格式，返回标准格式 YYYY-MM-DD"""
        date_str = str(date_str).strip()
        
        # 如果已经是标准格式，直接返回
        try:
            datetime.strptime(date_str, '%Y-%m-%d')
            return date_str
        except:
            pass
        
        # 0. 处理带时间的格式：2026-03-16 09:47:44、2026/03/16 09:47:44、2026.03.16 09:47:44
        time_formats = [
            '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%Y.%m.%d %H:%M:%S',
            '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M', '%Y.%m.%d %H:%M'
        ]
        
        for fmt in time_formats:
            try:
                parsed_datetime = datetime.strptime(date_str, fmt)
                return parsed_datetime.strftime('%Y-%m-%d')
            except:
                continue
        
        # 1. 处理斜杠分隔格式：2026/3/2、2026/03/02、3/13、3/14
        if '/' in date_str:
            parts = date_str.split('/')
            if len(parts) == 3:
                # 格式：2026/3/2 或 2026/03/02
                year = int(parts[0])
                month = int(parts[1])
                day = int(parts[2])
                return f"{year:04d}-{month:02d}-{day:02d}"
            elif len(parts) == 2:
                # 格式：3/13、3/14（自动识别今年年份）
                current_year = datetime.now().year
                month = int(parts[0])
                day = int(parts[1])
                return f"{current_year:04d}-{month:02d}-{day:02d}"
        
        # 2. 处理点分隔格式：3.13、3.14、3.15（自动识别今年年份）
        elif '.' in date_str:
            parts = date_str.split('.')
            if len(parts) == 2:
                current_year = datetime.now().year
                month = int(parts[0])
                day = int(parts[1])
                return f"{current_year:04d}-{month:02d}-{day:02d}"
            elif len(parts) == 3:
                # 格式：2026.3.2 或 2026.03.02
                year = int(parts[0])
                month = int(parts[1])
                day = int(parts[2])
                return f"{year:04d}-{month:02d}-{day:02d}"
        
        # 3. 处理横杠分隔格式：3-13、3-14、2026-3-2
        elif '-' in date_str:
            parts = date_str.split('-')
            if len(parts) == 2:
                # 格式：3-13、3-14（自动识别今年年份）
                current_year = datetime.now().year
                month = int(parts[0])
                day = int(parts[1])
                return f"{current_year:04d}-{month:02d}-{day:02d}"
            elif len(parts) == 3:
                # 格式：2026-3-2 或 2026-03-02
                year = int(parts[0])
                month = int(parts[1])
                day = int(parts[2])
                return f"{year:04d}-{month:02d}-{day:02d}"
        
        # 4. 处理中文格式：2026年3月2日、3月13日、3月14日
        if '年' in date_str and '月' in date_str and '日' in date_str:
            import re
            match = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日', date_str)
            if match:
                year = int(match.group(1))
                month = int(match.group(2))
                day = int(match.group(3))
                return f"{year:04d}-{month:02d}-{day:02d}"
            
            match = re.search(r'(\d{1,2})月(\d{1,2})日', date_str)
            if match:
                current_year = datetime.now().year
                month = int(match.group(1))
                day = int(match.group(2))
                return f"{current_year:04d}-{month:02d}-{day:02d}"
        
        # 5. 处理无分隔符格式：20260302、0302（自动识别今年年份）
        if date_str.isdigit():
            if len(date_str) == 8:
                # 格式：20260302
                year = int(date_str[:4])
                month = int(date_str[4:6])
                day = int(date_str[6:8])
                return f"{year:04d}-{month:02d}-{day:02d}"
            elif len(date_str) == 4:
                # 格式：0302（3月2日）、0313（3月13日）
                current_year = datetime.now().year
                month = int(date_str[:2])
                day = int(date_str[2:4])
                return f"{current_year:04d}-{month:02d}-{day:02d}"
        
        # 6. 尝试常见日期格式解析
        common_formats = [
            '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y',
            '%Y-%m-%d', '%m-%d-%Y', '%d-%m-%Y',
            '%Y.%m.%d', '%m.%d.%Y', '%d.%m.%Y',
            '%Y年%m月%d日', '%m月%d日',
            '%Y%m%d'
        ]
        
        for fmt in common_formats:
            try:
                parsed_date = datetime.strptime(date_str, fmt)
                return parsed_date.strftime('%Y-%m-%d')
            except:
                continue
        
        # 如果所有格式都解析失败，返回当前日期
        return self.get_current_date()

    def add_record(self):
        """添加记录"""
        store_id = self.store_combo.currentData()
        if store_id is None:
            QMessageBox.warning(self, "警告", "请选择店铺！")
            return
        order_no = self.order_no_edit.text().strip()
        if not order_no:
            QMessageBox.warning(self, "警告", "订单号不能为空！")
            return
        reason = self.reason_combo.currentText()
        if not reason:
            QMessageBox.warning(self, "警告", "请选择退款原因！")
            return
        try:
            refund_amount = float(self.refund_amount_edit.text())
        except ValueError:
            QMessageBox.warning(self, "警告", "退款金额必须为有效数字！")
            return
        cancel = self.cancel_check.isChecked()
        compensate = self.compensate_check.isChecked()
        comp_amount = 0.0
        if compensate:
            try:
                comp_amount = float(self.comp_amount_edit.text()) if self.comp_amount_edit.text() else 0.0
            except ValueError:
                QMessageBox.warning(self, "警告", "补偿金额必须为有效数字！")
                return
        
        # 驳回相关字段
        reject = self.reject_check.isChecked()
        reject_result = ""
        if reject:
            reject_result = self.reject_result_combo.currentText()
        
        notes = self.notes_edit.text().strip()
        
        record_date = self.get_current_date()

        # 订单号重复检查
        existing = self.db.get_record_by_order_no(order_no)
        if existing:
            QMessageBox.warning(self, "警告", f"订单号 {order_no} 已存在，无法重复添加！")
            return

        self.db.add_record(store_id, order_no, reason, refund_amount, cancel, compensate, comp_amount, reject, reject_result, notes, record_date)
        self.show_tooltip("已添加", "rgba(76, 175, 80, 0.95)", 1000)  # 绿色气泡显示1秒
        self.clear_input()
        self.load_table_data()

    def update_record(self):
        """更新记录"""
        if self.current_record_id is None:
            QMessageBox.warning(self, "警告", "请先在表格中选中要更新的记录！")
            return
        store_id = self.store_combo.currentData()
        if store_id is None:
            QMessageBox.warning(self, "警告", "请选择店铺！")
            return
        order_no = self.order_no_edit.text().strip()
        if not order_no:
            QMessageBox.warning(self, "警告", "订单号不能为空！")
            return
        reason = self.reason_combo.currentText()
        if not reason:
            QMessageBox.warning(self, "警告", "请选择退款原因！")
            return
        try:
            refund_amount = float(self.refund_amount_edit.text())
        except ValueError:
            QMessageBox.warning(self, "警告", "退款金额必须为有效数字！")
            return
        cancel = self.cancel_check.isChecked()
        compensate = self.compensate_check.isChecked()
        comp_amount = 0.0
        if compensate:
            try:
                comp_amount = float(self.comp_amount_edit.text()) if self.comp_amount_edit.text() else 0.0
            except ValueError:
                QMessageBox.warning(self, "警告", "补偿金额必须为有效数字！")
                return
        # 驳回相关字段
        reject = self.reject_check.isChecked()
        reject_result = ""
        if reject:
            reject_result = self.reject_result_combo.currentText()
        
        notes = self.notes_edit.text().strip()
        
        record_date = self.get_current_date()

        self.db.update_record(self.current_record_id, store_id, order_no, reason, refund_amount, cancel, compensate, comp_amount, reject, reject_result, notes, record_date)
        self.show_tooltip("已更新", "rgba(76, 175, 80, 0.95)", 1000)  # 绿色气泡显示1秒
        
        # 不清空输入区域，保持当前记录显示
        # 强制刷新表格数据，让用户看到更新效果
        self.load_table_data(force_reload=True)
        
        # 重新选中当前记录，让用户看到更新后的状态
        self._select_current_record_after_update()

    def delete_record(self):
        """删除选中的记录（支持多选删除）"""
        # 获取所有选中的行
        selected_rows = self.table.selectionModel().selectedRows()
        
        if not selected_rows:
            QMessageBox.warning(self, "警告", "请先在表格中选中要删除的记录！")
            return
        
        # 获取选中行的记录ID
        record_ids = []
        for index in selected_rows:
            row = index.row()
            record_id = self.get_record_id_from_row(row)
            if record_id:
                record_ids.append(record_id)
        
        if not record_ids:
            QMessageBox.warning(self, "警告", "无法获取选中记录的ID！")
            return
        
        # 确认删除对话框
        if len(record_ids) == 1:
            message = "确定要删除这条记录吗？"
        else:
            message = f"确定要删除选中的 {len(record_ids)} 条记录吗？"
            
        reply = QMessageBox.question(self, "确认删除", message,
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            # 批量删除记录
            success_count = 0
            failed_ids = []
            
            for record_id in record_ids:
                try:
                    if self.db.delete_record(record_id):
                        success_count += 1
                    else:
                        failed_ids.append(record_id)
                except Exception as e:
                    print(f"删除记录 {record_id} 时出错: {e}")
                    failed_ids.append(record_id)
            
            if success_count > 0:
                if success_count == 1:
                    QMessageBox.information(self, "成功", "记录已删除！")
                else:
                    QMessageBox.information(self, "成功", f"已成功删除 {success_count} 条记录！")
                
                # 清除输入并刷新表格
                self.clear_input()
                # 强制刷新表格数据（确保删除后立即消失）
                self.load_table_data(force_reload=True)
                # 强制刷新表格显示
                self.table.viewport().update()
                
                # 如果有失败的删除，显示警告
                if failed_ids:
                    QMessageBox.warning(self, "部分失败", f"成功删除 {success_count} 条记录，但 {len(failed_ids)} 条记录删除失败！")
            else:
                QMessageBox.warning(self, "错误", "所有记录删除失败！")

    def refund_amount_mouse_press(self, event):
        """退款金额输入框鼠标点击事件 - 只在有内容时自动全选"""
        # 只有当输入框有内容时才自动全选
        if self.refund_amount_edit.text():
            self.refund_amount_edit.selectAll()
        # 调用原始的鼠标点击事件
        QLineEdit.mousePressEvent(self.refund_amount_edit, event)

    def order_no_mouse_press(self, event):
        """订单号输入框鼠标点击事件 - 自动全选文本"""
        self.order_no_edit.selectAll()
        # 调用原始的鼠标点击事件
        QLineEdit.mousePressEvent(self.order_no_edit, event)

    def comp_amount_mouse_press(self, event):
        """补偿金额输入框鼠标点击事件 - 自动全选文本"""
        self.comp_amount_edit.selectAll()
        # 调用原始的鼠标点击事件
        QLineEdit.mousePressEvent(self.comp_amount_edit, event)

    def search_order_mouse_press(self, event):
        """搜索订单号输入框鼠标点击事件 - 自动全选文本"""
        self.search_order_edit.selectAll()
        # 调用原始的鼠标点击事件
        QLineEdit.mousePressEvent(self.search_order_edit, event)

    def clear_input(self):
        """清空输入区域"""
        self.store_combo.setCurrentIndex(0)
        self.order_no_edit.clear()
        self.reason_combo.setCurrentIndex(0)
        self.refund_amount_edit.clear()
        self.cancel_check.setChecked(False)
        self.compensate_check.setChecked(False)
        self.comp_amount_edit.clear()
        self.current_record_id = None
        self.table.clearSelection()

    def get_filtered_records(self):
        """获取当前筛选条件下的记录（与表格显示的数据相同）"""
        order_no = self.search_order_edit.text()
        
        # 处理退款原因筛选（支持多选）
        reason = self.search_reason_combo.checkedItems()
        if not reason:  # 如果没有选择任何原因，显示全部
            reason = "全部"
        
        cancel = self.search_cancel_combo.currentText()
        compensate = self.search_compensate_combo.currentText()
        reject = self.search_reject_combo.currentText()
        reject_result = self.search_reject_result_combo.currentText()
        store_name = self.search_store_combo.currentText()
        start_date = self.start_date_edit.date().toString("yyyy-MM-dd")
        end_date = self.end_date_edit.date().toString("yyyy-MM-dd")

        return self.db.search_records(order_no, reason, cancel, compensate, reject, reject_result, start_date, end_date, store_name)

    def load_table_data(self, force_reload=False):
        """加载表格数据（根据筛选条件）"""
        # 如果强制重新加载，忽略缓存
        if force_reload:
            self._cached_records = None
            self._last_search_params = None
        
        # 性能优化：缓存检查 - 如果搜索参数相同且数据已缓存，直接使用缓存
        current_params = self._get_current_search_params()
        if self._last_search_params == current_params and self._cached_records is not None:
            records = self._cached_records
        else:
            # 需要重新查询数据库
            records = self.get_filtered_records()
            # 性能优化：缓存搜索结果
            self._cached_records = records
            self._last_search_params = current_params
        
        # 在加载数据时暂时断开cellChanged信号，防止误触发
        try:
            self.table.cellChanged.disconnect(self.on_cell_changed)
        except TypeError:
            # 如果信号还没有连接，忽略错误
            pass
        
        # 获取筛选参数用于调试标签
        order_no = self.search_order_edit.text()
        reason = self.search_reason_combo.checkedItems()
        if not reason:
            reason = "全部"
        store_name = self.search_store_combo.currentText()
        
        # 更新调试标签显示当前筛选结果
        self.update_debug_label(len(records), order_no, reason, store_name)

        # 性能优化：增量更新 - 只更新变化的行
        current_row_count = self.table.rowCount()
        new_row_count = len(records)
        
        # 设置新的行数
        if new_row_count != current_row_count:
            self.table.setRowCount(new_row_count)
        
        # 性能优化：批量更新表格数据
        for row, rec in enumerate(records):
            # 获取店铺颜色
            store_color = self.db.get_store_color(rec['store_name'])
            
            # 性能优化：检查是否需要更新该行
            if self._should_update_row(row, rec):
                # 店铺名称
                store_item = QTableWidgetItem(rec['store_name'])
                if store_color:
                    store_item.setBackground(QColor(store_color))
                self.table.setItem(row, 0, store_item)
                
                # 订单号
                order_item = QTableWidgetItem(rec['order_no'])
                if store_color:
                    order_item.setBackground(QColor(store_color))
                self.table.setItem(row, 1, order_item)
                
                # 退款原因
                reason_item = QTableWidgetItem(rec['reason'])
                if store_color:
                    reason_item.setBackground(QColor(store_color))
                self.table.setItem(row, 2, reason_item)
                
                # 退款金额
                amount_item = QTableWidgetItem(f"¥{rec['refund_amount']:.2f}")
                amount_item.setTextAlignment(Qt.AlignCenter)
                if store_color:
                    amount_item.setBackground(QColor(store_color))
                self.table.setItem(row, 3, amount_item)
                
                # 撤销
                cancel_text = "是" if rec['cancel'] else "否"
                cancel_item = QTableWidgetItem(cancel_text)
                # 单个单元格变色：是=绿色背景+白色文字，否=红色背景+白色文字
                if rec['cancel']:
                    cancel_item.setBackground(QColor("#4CAF50"))  # 绿色背景
                    cancel_item.setForeground(QColor("white"))     # 白色文字
                else:
                    cancel_item.setBackground(QColor("#F44336"))  # 红色背景
                    cancel_item.setForeground(QColor("white"))     # 白色文字
                cancel_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row, 4, cancel_item)
                
                # 打款补偿
                comp_text = "是" if rec['compensate'] else "否"
                comp_item = QTableWidgetItem(comp_text)
                # 单个单元格变色：是=绿色背景+白色文字，否=红色背景+白色文字
                if rec['compensate']:
                    comp_item.setBackground(QColor("#4CAF50"))  # 绿色背景
                    comp_item.setForeground(QColor("white"))     # 白色文字
                else:
                    comp_item.setBackground(QColor("#F44336"))  # 红色背景
                    comp_item.setForeground(QColor("white"))     # 白色文字
                comp_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row, 5, comp_item)
                
                # 补偿金额
                comp_amount_item = QTableWidgetItem(f"¥{rec['comp_amount']:.2f}")
                comp_amount_item.setTextAlignment(Qt.AlignCenter)
                if store_color:
                    comp_amount_item.setBackground(QColor(store_color))
                self.table.setItem(row, 6, comp_amount_item)
                
                # 驳回
                reject_text = "是" if rec['reject'] else "否"
                reject_item = QTableWidgetItem(reject_text)
                # 单个单元格变色：是=绿色背景+白色文字，否=红色背景+白色文字
                if rec['reject']:
                    reject_item.setBackground(QColor("#4CAF50"))  # 绿色背景
                    reject_item.setForeground(QColor("white"))     # 白色文字
                else:
                    reject_item.setBackground(QColor("#F44336"))  # 红色背景
                    reject_item.setForeground(QColor("white"))     # 白色文字
                reject_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row, 7, reject_item)
                
                # 驳回结果
                reject_result_item = QTableWidgetItem(rec['reject_result'])
                reject_result_item.setTextAlignment(Qt.AlignCenter)
                if store_color:
                    reject_result_item.setBackground(QColor(store_color))
                self.table.setItem(row, 8, reject_result_item)
                
                # 登记日期
                date_item = QTableWidgetItem(rec['record_date'])
                if store_color:
                    date_item.setBackground(QColor(store_color))
                self.table.setItem(row, 9, date_item)
                
                # 备注
                notes_item = QTableWidgetItem(rec['notes'])
                if store_color:
                    notes_item.setBackground(QColor(store_color))
                self.table.setItem(row, 10, notes_item)

            # 高亮刚导入的订单（覆盖店铺颜色）
            if rec['order_no'] in self.highlighted_orders:
                for col in range(11):
                    if self.table.item(row, col):
                        self.table.item(row, col).setBackground(QColor("#FFD700"))  # 金色高亮
        
        # 数据加载完成后重新连接cellChanged信号
        self.table.cellChanged.connect(self.on_cell_changed)
        
        # 性能优化：合并统计更新，避免重复计算
        self._update_all_statistics(records)
        
        # 清除高亮标记（高亮持续到下次加载，但我们保留一个标记，下次加载时会重新根据集合高亮，直到用户点击表格）
        # 注意：用户点击表格时清除高亮，通过table的itemClicked信号实现

    def update_statusbar(self, records):
        """更新状态栏统计"""
        total = len(records)
        # 退款金额：只计算未撤销且（未驳回或驳回失败）的订单
        total_refund = sum(r['refund_amount'] for r in records 
                          if not r['cancel'] and (not r['reject'] or r['reject_result'] == '失败'))
        # 补偿总额：只要打款状态为"是"（已打款）的都要计算补偿金额
        total_comp = sum(r['comp_amount'] for r in records if r['compensate'])
        # 总金额：退款金额 + 补偿金额
        total_amount = total_refund + total_comp
        cancel_count = sum(1 for r in records if r['cancel'])
        # 计算驳回相关统计
        reject_count = sum(1 for r in records if r['reject'])
        reject_success_count = sum(1 for r in records if r['reject'] and r['reject_result'] == '成功')
        reject_fail_count = sum(1 for r in records if r['reject'] and r['reject_result'] == '失败')
        
        self.status_bar.showMessage(
            f"记录总数: {total} | 退款总额: ¥{total_refund:,.2f} | 补偿总额: ¥{total_comp:,.2f} | "
            f"总金额: ¥{total_amount:,.2f} | 撤销订单: {cancel_count}单 | 驳回: {reject_count}单(成功:{reject_success_count}/失败:{reject_fail_count})"
        )

    def on_search_changed(self):
        """搜索条件变化时自动搜索（实时搜索）"""
        # 性能优化：复用已有的定时器，避免重复创建
        self._search_timer.stop()
        
        def update_data():
            # 性能优化：合并数据库查询，一次搜索获取所有数据
            self.load_table_data()
            # 不再单独调用 update_store_stats_display()，因为 load_table_data() 中已经包含统计更新
        
        # 性能优化：确保定时器连接正确
        try:
            self._search_timer.timeout.disconnect()
        except:
            pass
        self._search_timer.timeout.connect(update_data)
        self._search_timer.start(800)

    def _get_current_search_params(self):
        """获取当前搜索参数（用于缓存检查）"""
        return (
            self.search_order_edit.text(),
            tuple(self.search_reason_combo.checkedItems()),
            self.search_cancel_combo.currentText(),
            self.search_compensate_combo.currentText(),
            self.search_reject_combo.currentText(),
            self.search_reject_result_combo.currentText(),
            self.search_store_combo.currentText(),
            self.start_date_edit.date().toString("yyyy-MM-dd"),
            self.end_date_edit.date().toString("yyyy-MM-dd")
        )

    def _should_update_row(self, row, record):
        """检查是否需要更新指定行（增量更新优化）"""
        # 如果行数变化，需要更新所有行
        if row >= self.table.rowCount():
            return True
        
        # 检查当前行数据是否与记录匹配
        current_store = self.table.item(row, 0).text() if self.table.item(row, 0) else ""
        current_order = self.table.item(row, 1).text() if self.table.item(row, 1) else ""
        
        # 如果店铺名称或订单号不匹配，需要更新
        if current_store != record['store_name'] or current_order != record['order_no']:
            return True
        
        # 检查状态字段是否变化（撤销、打款补偿、驳回）
        current_cancel = self.table.item(row, 4).text() if self.table.item(row, 4) else ""
        current_compensate = self.table.item(row, 5).text() if self.table.item(row, 5) else ""
        current_reject = self.table.item(row, 7).text() if self.table.item(row, 7) else ""
        
        # 如果状态字段变化，需要更新（确保颜色实时变化）
        expected_cancel = "是" if record['cancel'] else "否"
        expected_compensate = "是" if record['compensate'] else "否"
        expected_reject = "是" if record['reject'] else "否"
        
        if (current_cancel != expected_cancel or 
            current_compensate != expected_compensate or 
            current_reject != expected_reject):
            return True
        
        return False

    def _update_all_statistics(self, records):
        """合并更新所有统计信息（避免重复计算）"""
        # 更新状态栏统计
        self.update_statusbar(records)
        # 更新左下角总金额显示
        self.update_total_amount_display()
        # 更新店铺统计信息显示
        self.update_store_stats_display()

    def reset_search(self):
        """重置搜索条件"""
        self.search_order_edit.clear()
        self.search_store_combo.setCurrentIndex(0)  # 全部
        self.search_reason_combo.clearChecked()  # 清空多选状态
        self.search_cancel_combo.setCurrentIndex(0)  # 全部
        self.search_compensate_combo.setCurrentIndex(0)  # 全部
        self.search_reject_combo.setCurrentIndex(0)  # 全部
        self.search_reject_result_combo.setCurrentIndex(0)  # 全部
        today = QDate.currentDate()
        self.start_date_edit.setDate(today)
        self.end_date_edit.setDate(today)
        self.load_table_data()

    def show_all_records(self):
        """显示全部记录（清除所有筛选条件，强制重新加载）"""
        # 清除所有筛选条件
        self.search_order_edit.clear()
        self.search_store_combo.setCurrentIndex(0)  # 全部
        self.search_reason_combo.clearChecked()  # 清空多选状态
        self.search_cancel_combo.setCurrentIndex(0)  # 全部
        self.search_compensate_combo.setCurrentIndex(0)  # 全部
        self.search_reject_combo.setCurrentIndex(0)  # 全部
        self.search_reject_result_combo.setCurrentIndex(0)  # 全部
        
        # 设置日期为所有日期
        self.start_date_edit.setDate(QDate(2000, 1, 1))  # 很早的日期
        self.end_date_edit.setDate(QDate(2100, 12, 31))  # 很晚的日期
        
        # 强制重新加载所有数据
        self.load_table_data(force_reload=True)
        
        # 显示提示信息
        total_count = self.table.rowCount()
        QMessageBox.information(self, "显示全部", 
                               f"✅ 已显示全部记录！\n"
                               f"当前显示 {total_count} 条记录。")

    def set_quick_date(self, days):
        """快捷日期设置（近7天和近30天不包括今天）"""
        today = QDate.currentDate()
        if days == 0:  # 今天
            start = today
            end = today
        elif days == 1:  # 昨天
            start = today.addDays(-1)
            end = today.addDays(-1)
        else:
            # 近7天和近30天不包括今天，只计算完整一天的数据
            start = today.addDays(-days)  # 从昨天往前推days-1天
            end = today.addDays(-1)       # 到昨天为止
        self.start_date_edit.setDate(start)
        self.end_date_edit.setDate(end)
        self.load_table_data()

    def previous_day(self):
        """前一天：将当前日期范围往前移动一天"""
        current_start = self.start_date_edit.date()
        current_end = self.end_date_edit.date()
        
        # 如果开始日期和结束日期相同（单天选择）
        if current_start == current_end:
            new_date = current_start.addDays(-1)
            self.start_date_edit.setDate(new_date)
            self.end_date_edit.setDate(new_date)
        else:
            # 如果是多天选择，整体往前移动一天
            new_start = current_start.addDays(-1)
            new_end = current_end.addDays(-1)
            self.start_date_edit.setDate(new_start)
            self.end_date_edit.setDate(new_end)
        
        self.load_table_data()

    def next_day(self):
        """后一天：将当前日期范围往后移动一天"""
        current_start = self.start_date_edit.date()
        current_end = self.end_date_edit.date()
        
        # 如果开始日期和结束日期相同（单天选择）
        if current_start == current_end:
            new_date = current_start.addDays(1)
            # 检查是否超过今天
            today = QDate.currentDate()
            if new_date > today:
                new_date = today
            self.start_date_edit.setDate(new_date)
            self.end_date_edit.setDate(new_date)
        else:
            # 如果是多天选择，整体往后移动一天
            new_start = current_start.addDays(1)
            new_end = current_end.addDays(1)
            # 检查是否超过今天
            today = QDate.currentDate()
            if new_end > today:
                new_end = today
                new_start = new_end.addDays(-(current_end.daysTo(current_start)))
            self.start_date_edit.setDate(new_start)
            self.end_date_edit.setDate(new_end)
        
        self.load_table_data()

    def on_item_double_clicked(self, item):
        """双击表格项：根据列类型执行不同操作"""
        try:
            row = item.row()
            column = item.column()
            
            # 根据列类型执行不同操作
            if column == 0:  # 店铺名称列：录入信息
                self.load_record_to_input(row)
            elif column == 1:  # 订单号列：复制订单号
                self.copy_order_no(row)
            elif column == 2:  # 退款原因列：无操作
                pass
            elif column == 3:  # 退款金额列：直接编辑
                self.table.editItem(item)
            elif column == 4:  # 撤销列：双击切换
                self.toggle_status_field(row, column)
            elif column == 5:  # 打款补偿列：双击切换
                self.toggle_status_field(row, column)
            elif column == 6:  # 补偿金额列：条件编辑
                if self.table.item(row, 5).text() == "是":  # 只有打款补偿为"是"时才能编辑
                    self.table.editItem(item)
            elif column == 7:  # 驳回列：双击切换
                self.toggle_status_field(row, column)
            elif column == 8:  # 驳回结果列：条件下拉框选择
                if self.table.item(row, 7).text() == "是":  # 只有驳回为"是"时才能选择
                    self.show_reject_result_dropdown(row, column)
            elif column == 9:  # 登记日期列：无操作
                pass
            elif column == 10:  # 备注列：直接编辑（现在在第10列）
                self.table.editItem(item)
        except Exception as e:
            # 捕获所有异常，防止程序崩溃
            QMessageBox.warning(self, "操作错误", f"双击操作失败：{str(e)}")
        
    def load_record_to_input(self, row):
        """将选中行的数据录入到输入框（只有双击店铺名称列时调用）"""
        # 安全检查：确保行号有效
        if row < 0 or row >= self.table.rowCount():
            return
            
        # 安全检查：只检查必要的列（前6列必须有数据，后4列可以为空）
        required_columns = [0, 1, 2, 3, 4, 5]  # 店铺名称、订单号、退款原因、退款金额、撤销、打款补偿
        for col in required_columns:
            if not self.table.item(row, col):
                QMessageBox.warning(self, "错误", f"第{col+1}列数据缺失，无法加载")
                return
        
        # 获取选中行的数据
        store_name = self.table.item(row, 0).text()
        order_no = self.table.item(row, 1).text()
        reason = self.table.item(row, 2).text()
        refund_amount_text = self.table.item(row, 3).text()
        cancel_text = self.table.item(row, 4).text()
        compensate_text = self.table.item(row, 5).text()
        comp_amount_text = self.table.item(row, 6).text()
        reject_text = self.table.item(row, 7).text()
        reject_result_text = self.table.item(row, 8).text()
        notes_text = self.table.item(row, 10).text()  # 修复：备注列索引应该是10，不是9
        
        # 解析退款金额（去掉¥符号）
        try:
            refund_amount = float(refund_amount_text.replace('¥', '').strip())
        except:
            refund_amount = 0.0
            
        # 解析补偿金额
        try:
            comp_amount = float(comp_amount_text.replace('¥', '').strip()) if comp_amount_text else 0.0
        except:
            comp_amount = 0.0
            
        # 设置店铺
        store_index = self.store_combo.findText(store_name)
        if store_index >= 0:
            self.store_combo.setCurrentIndex(store_index)
            
        # 设置订单号
        self.order_no_edit.setText(order_no)
        
        # 设置退款原因
        reason_index = self.reason_combo.findText(reason)
        if reason_index >= 0:
            self.reason_combo.setCurrentIndex(reason_index)
        else:
            self.reason_combo.setCurrentIndex(0)
            
        # 设置退款金额
        self.refund_amount_edit.setText(f"{refund_amount:.2f}")
        
        # 设置撤销状态
        self.cancel_check.setChecked(cancel_text == "是")
        
        # 设置补偿状态和金额
        self.compensate_check.setChecked(compensate_text == "是")
        self.comp_amount_edit.setText(f"{comp_amount:.2f}" if comp_amount > 0 else "")
        
        # 设置驳回状态和结果
        self.reject_check.setChecked(reject_text == "是")
        reject_result_index = self.reject_result_combo.findText(reject_result_text)
        if reject_result_index >= 0:
            self.reject_result_combo.setCurrentIndex(reject_result_index)
        else:
            self.reject_result_combo.setCurrentIndex(0)
            
        # 设置备注
        self.notes_edit.setText(notes_text)
        
        # 设置当前记录ID
        rec = self.db.get_record_by_order_no(order_no)
        if rec:
            self.current_record_id = rec['id']
        else:
            self.current_record_id = None

    def on_cell_changed(self, row, column):
        """表格单元格编辑完成时触发"""
        # 防止递归调用
        self.table.cellChanged.disconnect(self.on_cell_changed)
        
        try:
            item = self.table.item(row, column)
            if not item:
                return
                
            # 获取记录ID
            record_id = self.get_record_id_from_row(row)
            if not record_id:
                return
            
            # 根据列索引处理不同的字段
            if column == 3:  # 退款金额列
                self.update_refund_amount(record_id, item.text())
            elif column == 6:  # 补偿金额列
                self.update_comp_amount(record_id, item.text())
            elif column in [4, 5, 7]:  # 撤销、打款补偿、驳回状态列
                # 处理状态字段编辑：自动标准化输入
                text = item.text().strip()
                
                # 自动标准化输入
                if text.lower() in ['是', 'true', '1', 'yes', 'y', 't']:
                    item.setText("是")
                    self.update_status_field(record_id, column, "是")
                elif text.lower() in ['否', 'false', '0', 'no', 'n', 'f']:
                    item.setText("否")
                    self.update_status_field(record_id, column, "否")
                else:
                    # 无效输入，恢复原值
                    rec = self.db.get_record_by_id(record_id)
                    if rec:
                        if column == 4:  # 撤销
                            original_value = "是" if rec['cancel'] else "否"
                        elif column == 5:  # 打款补偿
                            original_value = "是" if rec['compensate'] else "否"
                        elif column == 7:  # 驳回
                            original_value = "是" if rec['reject'] else "否"
                        item.setText(original_value)
                        QMessageBox.warning(self, "输入错误", "请输入'是'或'否'")
                
        finally:
            # 重新连接信号
            self.table.cellChanged.connect(self.on_cell_changed)
            
            # 强制刷新表格数据，更新颜色和统计信息
            self.load_table_data()

    def get_record_id_from_row(self, row):
        """根据行号获取记录ID"""
        order_no_item = self.table.item(row, 1)  # 订单号列
        if not order_no_item:
            return None
            
        order_no = order_no_item.text()
        records = self.db.get_record_by_order_no(order_no)
        if records:
            return records['id']
        return None

    def toggle_status_field(self, row, column):
        """双击切换状态字段（撤销、打款补偿、驳回）"""
        try:
            # 获取记录ID
            record_id = self.get_record_id_from_row(row)
            if not record_id:
                return
                
            # 获取当前记录信息
            rec = self.db.get_record_by_id(record_id)
            if not rec:
                return
                
            # 根据列索引确定要切换的字段
            if column == 4:  # 撤销列
                new_cancel = not rec['cancel']  # 切换状态
                self.db.update_record(
                    record_id, rec['store_id'], rec['order_no'], rec['reason'], 
                    rec['refund_amount'], new_cancel, rec['compensate'], rec['comp_amount'],
                    rec['reject'], rec['reject_result'], rec['notes'], rec['record_date']
                )
            elif column == 5:  # 打款补偿列
                new_compensate = not rec['compensate']  # 切换状态
                self.db.update_record(
                    record_id, rec['store_id'], rec['order_no'], rec['reason'], 
                    rec['refund_amount'], rec['cancel'], new_compensate, rec['comp_amount'],
                    rec['reject'], rec['reject_result'], rec['notes'], rec['record_date']
                )
            elif column == 7:  # 驳回列
                new_reject = not rec['reject']  # 切换状态
                self.db.update_record(
                    record_id, rec['store_id'], rec['order_no'], rec['reason'], 
                    rec['refund_amount'], rec['cancel'], rec['compensate'], rec['comp_amount'],
                    new_reject, rec['reject_result'], rec['notes'], rec['record_date']
                )
            
            # 强制刷新整个表格，忽略缓存
            self.load_table_data(force_reload=True)
            
        except Exception as e:
            # 如果出错，也强制刷新表格确保一致性
            self.load_table_data(force_reload=True)


    
    def _select_current_record_after_update(self):
        """更新记录后重新选中当前记录"""
        if self.current_record_id is None:
            return
            
        # 根据记录ID找到对应的行号
        for row in range(self.table.rowCount()):
            record_id = self.get_record_id_from_row(row)
            if record_id == self.current_record_id:
                # 选中该行
                self.table.selectRow(row)
                # 滚动到该行
                self.table.scrollToItem(self.table.item(row, 0))
                break
    
    def _update_statistics_only(self):
        """只更新统计信息，不刷新整个表格"""
        # 获取当前筛选条件下的记录
        records = self.get_filtered_records()
        # 更新状态栏统计
        self.update_statusbar(records)

    def update_status_field(self, record_id, column, value):
        """更新状态字段（撤销、打款补偿、驳回）"""
        # 获取当前记录信息
        rec = self.db.get_record_by_id(record_id)
        if not rec:
            return
            
        # 根据列索引确定要更新的字段
        if column == 4:  # 撤销列
            cancel = value.lower() in ['是', 'true', '1', 'yes']
            self.db.update_record(
                record_id, rec['store_id'], rec['order_no'], rec['reason'], 
                rec['refund_amount'], cancel, rec['compensate'], rec['comp_amount'],
                rec['reject'], rec['reject_result'], rec['notes'], rec['record_date']
            )
        elif column == 5:  # 打款补偿列
            compensate = value.lower() in ['是', 'true', '1', 'yes']
            self.db.update_record(
                record_id, rec['store_id'], rec['order_no'], rec['reason'], 
                rec['refund_amount'], rec['cancel'], compensate, rec['comp_amount'],
                rec['reject'], rec['reject_result'], rec['notes'], rec['record_date']
            )
        elif column == 7:  # 驳回列
            reject = value.lower() in ['是', 'true', '1', 'yes']
            self.db.update_record(
                record_id, rec['store_id'], rec['order_no'], rec['reason'], 
                rec['refund_amount'], rec['cancel'], rec['compensate'], rec['comp_amount'],
                reject, rec['reject_result'], rec['notes'], rec['record_date']
            )

    def update_refund_amount(self, record_id, amount_text):
        """更新退款金额"""
        try:
            # 提取数字部分
            amount = float(amount_text.replace('¥', '').strip())
            if self.db.update_refund_amount(record_id, amount):
                # 更新显示格式
                row = self.get_row_from_record_id(record_id)
                if row is not None:
                    item = self.table.item(row, 3)
                    item.setText(f"¥{amount:.2f}")
                # 只更新统计信息，不重新加载整个表格，避免无限循环
                self.update_statusbar(self.get_filtered_records())
                self.update_total_amount_display()
                self.update_store_stats_display()
                self.show_tooltip("退款金额已更新", "rgba(76, 175, 80, 0.95)", 1000)  # 绿色气泡显示1秒
        except ValueError:
            QMessageBox.warning(self, "错误", "请输入有效的金额数字")
            # 重新加载数据恢复原值
            self.load_table_data()

    def update_comp_amount(self, record_id, amount_text):
        """更新补偿金额"""
        try:
            # 提取数字部分
            amount = float(amount_text.replace('¥', '').strip())
            if self.db.update_comp_amount(record_id, amount):
                # 更新显示格式
                row = self.get_row_from_record_id(record_id)
                if row is not None:
                    item = self.table.item(row, 6)
                    item.setText(f"¥{amount:.2f}")
                    self.show_tooltip("补偿金额已更新", "rgba(76, 175, 80, 0.95)", 1000)  # 绿色气泡显示1秒
        except ValueError:
            QMessageBox.warning(self, "错误", "请输入有效的金额数字")
            # 重新加载数据恢复原值
            self.load_table_data()

    def get_row_from_record_id(self, record_id):
        """根据记录ID获取行号"""
        for row in range(self.table.rowCount()):
            current_id = self.get_record_id_from_row(row)
            if current_id == record_id:
                return row
        return None

    def on_item_clicked(self, item):
        """单击表格项：自动录入订单信息到输入框"""
        row = item.row()
        
        # 获取选中行的数据
        store_name = self.table.item(row, 0).text()
        order_no = self.table.item(row, 1).text()
        reason = self.table.item(row, 2).text()
        refund_amount_text = self.table.item(row, 3).text()
        cancel_text = self.table.item(row, 4).text()
        compensate_text = self.table.item(row, 5).text()
        comp_amount_text = self.table.item(row, 6).text()
        
        # 解析退款金额（去掉¥符号）
        try:
            refund_amount = float(refund_amount_text.replace('¥', '').strip())
        except:
            refund_amount = 0.0
            
        # 解析补偿金额
        try:
            comp_amount = float(comp_amount_text.replace('¥', '').strip()) if comp_amount_text else 0.0
        except:
            comp_amount = 0.0
            
        # 设置店铺
        store_index = self.store_combo.findText(store_name)
        if store_index >= 0:
            self.store_combo.setCurrentIndex(store_index)
            
        # 设置订单号
        self.order_no_edit.setText(order_no)
        
        # 设置退款原因
        reason_index = self.reason_combo.findText(reason)
        if reason_index >= 0:
            self.reason_combo.setCurrentIndex(reason_index)
        else:
            self.reason_combo.setCurrentIndex(0)
            
        # 设置退款金额
        self.refund_amount_edit.setText(f"{refund_amount:.2f}")
        
        # 设置撤销状态
        self.cancel_check.setChecked(cancel_text == "是")
        
        # 设置补偿状态和金额
        self.compensate_check.setChecked(compensate_text == "是")
        self.comp_amount_edit.setText(f"{comp_amount:.2f}" if comp_amount > 0 else "")
        
        # 记录当前记录ID
        rec = self.db.get_record_by_order_no(order_no)
        if rec:
            self.current_record_id = rec['id']
        else:
            self.current_record_id = None
            
        # 点击表格后清除高亮
        if self.highlighted_orders:
            self.highlighted_orders.clear()
            # 立即刷新表格显示，清除高亮
            self.load_table_data()

    # 双击功能已改为单击，此方法不再使用
    # def on_row_double_clicked(self, item):
    #     """双击行：填充到输入区"""
    #     row = item.row()
    #     store_name = self.table.item(row, 0).text()
    #     order_no = self.table.item(row, 1).text()
    #     reason = self.table.item(row, 2).text()
    #     refund_amount_text = self.table.item(row, 3).text().replace('¥', '').replace(',', '')
    #     try:
    #         refund_amount = float(refund_amount_text)
    #     except:
    #         refund_amount = 0.0
    #     cancel_text = self.table.item(row, 4).text()
    #     compensate_text = self.table.item(row, 5).text()
    #     comp_amount_text = self.table.item(row, 6).text().replace('¥', '').replace(',', '')
    #     try:
    #         comp_amount = float(comp_amount_text)
    #     except:
    #         comp_amount = 0.0

    #     # 设置店铺
    #     stores = self.db.get_stores()
    #     for idx, (sid, sname) in enumerate(stores):
    #         if sname == store_name:
    #             self.store_combo.setCurrentIndex(idx)
    #             break
    #     self.order_no_edit.setText(order_no)
    #     # 退款原因
    #     idx = self.reason_combo.findText(reason)
    #     if idx >= 0:
    #         self.reason_combo.setCurrentIndex(idx)
    #     else:
    #         # 如果原因不在列表中，添加并选中
    #         self.reason_combo.addItem(reason)
    #         self.reason_combo.setCurrentText(reason)
    #     self.refund_amount_edit.setText(str(refund_amount))
    #     self.cancel_check.setChecked(cancel_text == "是")
    #     self.compensate_check.setChecked(compensate_text == "是")
    #     if compensate_text == "是":
    #         self.comp_amount_edit.setEnabled(True)
    #         self.comp_amount_edit.setText(str(comp_amount) if comp_amount != 0 else "")
    #     else:
    #         self.comp_amount_edit.setEnabled(False)
    #         self.comp_amount_edit.clear()
    #     # 记录当前编辑的ID
    #     order_no = self.table.item(row, 1).text()
    #     rec = self.db.get_record_by_order_no(order_no)
    #     if rec:
    #         self.current_record_id = rec['id']
    #     else:
    #         self.current_record_id = None

    def show_context_menu(self, pos):
        """显示右键菜单"""
        item = self.table.itemAt(pos)
        
        # 创建自定义右键菜单
        menu = QMenu(self)
        
        if item is not None:
            # 如果点击了具体行，显示行操作菜单
            row = item.row()
            order_no = self.table.item(row, 1).text()
            store_name = self.table.item(row, 0).text()
            
            copy_order_action = QAction("复制订单号", self)
            copy_order_action.triggered.connect(lambda: self.copy_to_clipboard(order_no))
            copy_store_action = QAction("复制店铺名称", self)
            copy_store_action.triggered.connect(lambda: self.copy_to_clipboard(store_name))
            edit_action = QAction("编辑记录", self)
            edit_action.triggered.connect(lambda: self.on_item_clicked(self.table.item(row, 0)))
            delete_action = QAction("删除记录", self)
            delete_action.triggered.connect(self.delete_record)
            
            menu.addAction(copy_order_action)
            menu.addAction(copy_store_action)
            menu.addAction(edit_action)
            menu.addAction(delete_action)
            menu.addSeparator()
        
        # 添加全选当前筛选订单功能（无论是否点击具体行都显示）
        select_all_action = QAction("全选当前筛选订单", self)
        select_all_action.triggered.connect(self.select_all_filtered_orders)
        menu.addAction(select_all_action)
        
        menu.exec_(self.table.mapToGlobal(pos))

    def select_all_filtered_orders(self):
        """全选当前筛选出来的所有订单"""
        self.table.selectAll()
        selected_count = len(self.table.selectedItems()) // self.table.columnCount()
        self.show_tooltip(f"已选择 {selected_count} 条", "rgba(0, 120, 212, 0.95)", 1000)  # 蓝色气泡显示1秒





    def copy_to_clipboard(self, text):
        clipboard = QApplication.clipboard()
        clipboard.setText(text)
        self.show_tooltip("已复制", "rgba(76, 175, 80, 0.95)", 1000)  # 绿色气泡显示1秒

    def refresh_table_format(self):
        """刷新表格格式，清除导入后的高亮显示"""
        # 清除高亮订单集合
        if hasattr(self, 'highlighted_orders'):
            self.highlighted_orders.clear()
            print("[DEBUG] 已清除高亮订单集合")
        
        # 重新加载表格数据，清除高亮显示
        self.load_table_data()
        
        # 显示丝滑的气泡提示"已刷新"
        self.show_refresh_tooltip()

    # ---------------------------- 导入导出功能 ---------------------------------
    def export_excel(self):
        """导出当前表格数据到Excel"""
        # 获取当前表格中显示的数据（筛选后的）
        rows = self.table.rowCount()
        if rows == 0:
            QMessageBox.information(self, "提示", "没有数据可导出")
            return

        # 选择保存路径
        default_name = f"退款记录_{datetime.now().strftime('%Y%m%d')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(self, "导出Excel", default_name, "Excel文件 (*.xlsx)")
        if not file_path:
            return

        try:
            # 检查文件是否被占用
            import os
            if os.path.exists(file_path):
                try:
                    # 尝试以写入模式打开文件，如果被占用会抛出异常
                    with open(file_path, 'a', encoding='utf-8') as f:
                        pass
                except PermissionError:
                    QMessageBox.warning(self, "文件被占用", 
                                       f"文件 '{os.path.basename(file_path)}' 正在被其他程序使用！\n\n请先关闭该文件，然后重试。")
                    return
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "退款记录"

            # 表头 - 包含所有11列数据
            headers = ["店铺名称", "订单号", "退款原因", "退款金额", "撤销", "打款补偿", "补偿金额", "驳回", "驳回结果", "登记日期", "备注"]
            ws.append(headers)

            # 样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            center_alignment = Alignment(horizontal="center", vertical="center")

            # 应用表头样式
            for col in range(1, len(headers)+1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # 写入数据 - 导出所有11列
            for row_idx in range(rows):
                row_data = []
                for col in range(11):  # 改为11列
                    item = self.table.item(row_idx, col)
                    text = item.text() if item else ""
                    
                    # 处理金额格式
                    if col in [3, 6]:  # 退款金额和补偿金额列
                        # 去掉¥符号，保留数字
                        text = text.replace('¥', '').replace(',', '')
                    
                    # 处理日期格式 - 确保登记日期是准确日期
                    if col == 9:  # 登记日期列（第10列，索引为9）
                        # 如果日期显示有问题，尝试从数据库获取准确日期
                        record_id = self.get_record_id_from_row(row_idx)
                        if record_id:
                            record = self.db.get_record_by_id(record_id)
                            if record and record.get('record_date'):
                                text = record['record_date']  # 使用数据库中的准确日期
                    
                    row_data.append(text)
                ws.append(row_data)

            # 设置数据行样式（居中对齐，边框）
            for row_idx in range(2, rows+2):
                for col_idx in range(1, 12):  # 改为12列
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    
                    # 金额列格式为数字
                    if col_idx in [4, 7]:  # 退款金额和补偿金额列
                        try:
                            if cell.value:
                                cell.value = float(cell.value)
                        except:
                            pass
                    
                    # 日期列格式为日期
                    if col_idx == 10:  # 登记日期列
                        try:
                            if cell.value:
                                # 尝试解析日期格式
                                date_obj = datetime.strptime(cell.value, '%Y-%m-%d')
                                cell.value = date_obj
                                cell.number_format = 'YYYY-MM-DD'
                        except:
                            pass

            # 自动调整列宽
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[col_letter].width = adjusted_width

            wb.save(file_path)
            self.show_tooltip("导出成功", "rgba(76, 175, 80, 0.95)", 1500)  # 绿色气泡显示1.5秒
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败：{str(e)}")

    def fuzzy_match_column(self, headers, target_keywords):
        """模糊匹配列名：检查headers中是否包含目标关键词"""
        for header in headers:
            if header is None:
                continue
            header_str = str(header).strip()
            # 检查header是否包含所有目标关键词
            if all(keyword in header_str for keyword in target_keywords):
                return header_str
        return None

    def check_required_columns(self, headers, required_config):
        """检查必要列：支持模糊匹配"""
        missing_columns = []
        column_mapping = {}
        
        for col_config in required_config:
            if isinstance(col_config, str):
                # 简单字符串匹配
                if col_config not in headers:
                    missing_columns.append(col_config)
                else:
                    column_mapping[col_config] = col_config
            elif isinstance(col_config, dict):
                # 模糊匹配配置
                target_name = col_config['target']
                keywords = col_config['keywords']
                
                # 尝试精确匹配
                if target_name in headers:
                    column_mapping[target_name] = target_name
                    continue
                
                # 尝试模糊匹配
                matched_header = self.fuzzy_match_column(headers, keywords)
                if matched_header:
                    column_mapping[target_name] = matched_header
                else:
                    missing_columns.append(target_name)
        
        return missing_columns, column_mapping

    def import_excel(self):
        """导入Excel文件（智能模糊导入）"""
        file_path, _ = QFileDialog.getOpenFileName(self, "导入订单", "", "Excel文件 (*.xlsx *.xls)")
        if not file_path:
            return

        # 解析Excel
        data_rows = []
        column_mapping = {}
        try:
            if file_path.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                headers = [str(cell.value) if cell.value else "" for cell in ws[1]]
                
                # 显示表头识别信息
                header_info = "检测到的表头：\n" + "\n".join([f"{i+1}. {header}" for i, header in enumerate(headers)])
                QMessageBox.information(self, "表头识别", header_info)
                
                # 智能模糊匹配所有字段
                column_configs = [
                    {'target': '店铺名称', 'keywords': ['店铺', '名称', '店名', '门店'], 'required': False},
                    {'target': '订单号', 'keywords': ['订单', '编号', '单号', 'order'], 'required': True},
                    {'target': '退款原因', 'keywords': ['退款', '原因', '理由', '原因说明'], 'required': True},
                    {'target': '退款金额', 'keywords': ['退款', '金额', '钱', 'amount', '金额'], 'required': True},
                    {'target': '撤销', 'keywords': ['撤销', '取消', '撤单'], 'required': False},
                    {'target': '打款补偿', 'keywords': ['打款', '补偿', '赔付', '赔偿'], 'required': False},
                    {'target': '补偿金额', 'keywords': ['补偿', '金额', '赔款', '赔偿金额'], 'required': False},
                    {'target': '驳回', 'keywords': ['驳回', '拒绝', '不通过'], 'required': False},
                    {'target': '驳回结果', 'keywords': ['驳回', '结果', '处理结果'], 'required': False},
                    {'target': '备注', 'keywords': ['备注', '说明', '注释', 'note'], 'required': False},
                    {'target': '登记日期', 'keywords': ['日期', '时间', '登记', 'record', 'date'], 'required': False}
                ]
                
                # 智能匹配所有字段
                matched_columns = []
                for config in column_configs:
                    target_name = config['target']
                    keywords = config['keywords']
                    
                    # 尝试精确匹配
                    if target_name in headers:
                        column_mapping[target_name] = target_name
                        matched_columns.append(f"✅ {target_name} -> {target_name}")
                        continue
                    
                    # 尝试模糊匹配
                    matched_header = self.fuzzy_match_column(headers, keywords)
                    if matched_header:
                        column_mapping[target_name] = matched_header
                        matched_columns.append(f"✅ {target_name} -> {matched_header}")
                    else:
                        matched_columns.append(f"❌ {target_name} -> 未识别")
                
                # 显示匹配结果
                match_info = "字段匹配结果：\n" + "\n".join(matched_columns)
                QMessageBox.information(self, "字段匹配", match_info)
                
                # 检查必要列
                missing_required = []
                for config in column_configs:
                    if config['required'] and config['target'] not in column_mapping:
                        missing_required.append(config['target'])
                
                if missing_required:
                    QMessageBox.critical(self, "错误", f"缺少必要字段：{', '.join(missing_required)}")
                    return
                
                # 读取数据行，读取所有列（不仅仅是必要列）
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):  # 空行跳过
                        continue
                    row_dict = {}
                    for idx, val in enumerate(row):
                        if idx < len(headers):
                            header_name = headers[idx]
                            # 读取所有列，而不仅仅是必要列
                            row_dict[header_name] = val
                    data_rows.append(row_dict)
            elif file_path.endswith('.xls'):
                workbook = xlrd.open_workbook(file_path)
                sheet = workbook.sheet_by_index(0)
                headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
                
                # 检查必要列：根据搜索筛选区选择动态调整，支持模糊匹配
                current_search_store = self.search_store_combo.currentText()
                if current_search_store and current_search_store != "全部":
                    # 选择了具体店铺，店铺名称列可选
                    required_config = [
                        {'target': '订单号', 'keywords': ['订单']},
                        {'target': '退款原因', 'keywords': ['退款', '原因']},
                        {'target': '退款金额', 'keywords': ['退款', '金额']}
                    ]
                else:
                    # 选择了"全部"，店铺名称列为必要列
                    required_config = [
                        {'target': '店铺名称', 'keywords': ['店铺', '名称']},
                        {'target': '订单号', 'keywords': ['订单']},
                        {'target': '退款原因', 'keywords': ['退款', '原因']},
                        {'target': '退款金额', 'keywords': ['退款', '金额']}
                    ]
                
                # 检查必要列
                missing_columns, column_mapping = self.check_required_columns(headers, required_config)
                if missing_columns:
                    QMessageBox.critical(self, "错误", f"Excel缺少必要列：{', '.join(missing_columns)}")
                    return
                
                # 读取数据行，读取所有列（不仅仅是必要列）
                for row_idx in range(1, sheet.nrows):
                    row_dict = {}
                    for col_idx in range(sheet.ncols):
                        val = sheet.cell_value(row_idx, col_idx)
                        header_name = headers[col_idx]
                        # 读取所有列，而不仅仅是必要列
                        row_dict[header_name] = val
                    data_rows.append(row_dict)
            else:
                QMessageBox.critical(self, "错误", "不支持的文件格式")
                return
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取文件失败：{str(e)}")
            return

        if not data_rows:
            QMessageBox.information(self, "提示", "Excel中没有数据")
            return

        # 显示导入文件基本信息
        total_rows = len(data_rows)
        QMessageBox.information(self, "导入文件信息", 
                               f"Excel文件包含 {total_rows} 条数据\n\n"
                               f"开始导入处理...")

        # 处理导入
        success_count = 0
        overwrite_count = 0
        skip_count = 0
        fail_count = 0
        duplicate_count = 0
        self.highlighted_orders.clear()
        
        # 收集所有重复订单信息
        duplicate_orders = []  # 存储重复订单信息
        valid_rows = []  # 存储有效的行数据
        
        # 第一步：预处理所有数据，收集重复订单信息
        for row in data_rows:
            try:
                # 使用列映射获取正确的字段值
                store_name = ''
                if '店铺名称' in column_mapping:
                    actual_store_col = column_mapping['店铺名称']
                    store_name = str(row.get(actual_store_col, '')).strip()
                
                # 如果店铺名称为空，检查搜索筛选区是否选择了具体店铺
                if not store_name:
                    # 获取当前搜索筛选区选择的店铺
                    current_search_store = self.search_store_combo.currentText()
                    if current_search_store and current_search_store != "全部":
                        # 使用搜索筛选区选择的店铺名称
                        store_name = current_search_store
                    else:
                        # 没有选择具体店铺，跳过该行
                        fail_count += 1
                        continue
                
                # 使用列映射获取订单号
                order_no = ''
                if '订单号' in column_mapping:
                    actual_order_col = column_mapping['订单号']
                    order_no = str(row.get(actual_order_col, '')).strip()
                if not order_no:
                    fail_count += 1
                    continue
                
                # 使用列映射获取退款原因
                reason = ''
                if '退款原因' in column_mapping:
                    actual_reason_col = column_mapping['退款原因']
                    reason = str(row.get(actual_reason_col, '')).strip()
                if not reason:
                    fail_count += 1
                    continue
                
                # 定义品质退款原因列表（除了"其他"之外的所有原因）
                quality_reasons = ["商品腐败、变质、包装胀气等", "商品破损/压坏", "质量问题", 
                                  "大小/规格/重量等与商品描述不符", "品种/标签/图片/包装等与商品描述不符", "货物与描述不符"]
                
                # 如果导入的原因不在品质退款列表中，则归类为"其他"
                if reason not in quality_reasons and reason != "其他":
                    reason = "其他"
                
                # 使用列映射获取退款金额
                refund_amount = 0.0
                if '退款金额' in column_mapping:
                    actual_amount_col = column_mapping['退款金额']
                    refund_amount = row.get(actual_amount_col)
                try:
                    refund_amount = float(refund_amount)
                except:
                    fail_count += 1
                    continue
                
                # 可选字段 - 使用列映射获取
                compensate = '否'
                if '打款补偿' in column_mapping:
                    actual_compensate_col = column_mapping['打款补偿']
                    compensate = row.get(actual_compensate_col, '否')
                if isinstance(compensate, str):
                    compensate = compensate.strip() in ['是', 'True', 'true', '1', 'TRUE']
                else:
                    compensate = bool(compensate)
                
                # 调试输出打款补偿状态
                print(f"[DEBUG] 打款补偿字段值: {compensate} (原始值: {row.get(actual_compensate_col, '否') if '打款补偿' in column_mapping else '默认否'})")
                
                comp_amount = 0.0
                if '补偿金额' in column_mapping:
                    actual_comp_amount_col = column_mapping['补偿金额']
                    comp_amount = row.get(actual_comp_amount_col, 0)
                try:
                    comp_amount = float(comp_amount) if comp_amount else 0.0
                except:
                    comp_amount = 0.0
                
                # 撤销字段默认为否
                cancel = '否'
                if '撤销' in column_mapping:
                    actual_cancel_col = column_mapping['撤销']
                    cancel = row.get(actual_cancel_col, '否')
                if isinstance(cancel, str):
                    cancel = cancel.strip() in ['是', 'True', 'true', '1', 'TRUE']
                else:
                    cancel = bool(cancel)
                
                # 调试输出撤销状态
                print(f"[DEBUG] 撤销字段值: {cancel} (原始值: {row.get(actual_cancel_col, '否') if '撤销' in column_mapping else '默认否'})")
                
                # 驳回字段默认为否
                reject = '否'
                if '驳回' in column_mapping:
                    actual_reject_col = column_mapping['驳回']
                    reject = row.get(actual_reject_col, '否')
                if isinstance(reject, str):
                    reject = reject.strip() in ['是', 'True', 'true', '1']
                else:
                    reject = bool(reject)
                
                # 驳回结果字段：如果驳回为否，则设置为"无"；否则使用Excel中的值或默认""
                reject_result = ''
                if '驳回结果' in column_mapping:
                    actual_reject_result_col = column_mapping['驳回结果']
                    reject_result = row.get(actual_reject_result_col, '')
                if isinstance(reject_result, str):
                    reject_result = reject_result.strip()
                else:
                    reject_result = str(reject_result) if reject_result else ''
                
                # 如果驳回为否，则驳回结果强制设置为"无"
                if not reject:
                    reject_result = "无"
                
                # 备注字段默认为空
                notes = ''
                if '备注' in column_mapping:
                    actual_notes_col = column_mapping['备注']
                    notes = row.get(actual_notes_col, '')
                if isinstance(notes, str):
                    notes = notes.strip()
                else:
                    notes = str(notes) if notes else ''
                
                # 处理登记日期字段 - 支持多种表头名称，包括带时间的格式
                record_date = ''
                
                # 尝试多种可能的表头名称
                date_headers = ['登记日期', '登记时间', '日期', '时间', '创建日期', '创建时间']
                print(f"[DEBUG] 检查日期字段，当前行键值：{list(row.keys())}")
                for header in date_headers:
                    if header in row:
                        date_value = row.get(header, '')
                        if date_value:
                            print(f"[DEBUG] 找到日期字段 '{header}'，值：'{date_value}'")
                            # 尝试从带时间的字符串中提取日期部分
                            if isinstance(date_value, str):
                                # 首先检查是否是只有月份和日期的格式（如：3-16、3/16、3.16）
                                simple_date_pattern = r'^(\d{1,2})[-/.]?(\d{1,2})$'
                                import re
                                simple_match = re.match(simple_date_pattern, str(date_value).strip())
                                if simple_match:
                                    # 格式：3-16、3/16、3.16（只有月份和日期）
                                    current_year = datetime.now().year
                                    month = int(simple_match.group(1))
                                    day = int(simple_match.group(2))
                                    if 1 <= month <= 12 and 1 <= day <= 31:
                                        record_date = f"{current_year:04d}-{month:02d}-{day:02d}"
                                        print(f"[DEBUG] 简单日期格式匹配成功：{date_value} -> {record_date}")
                                
                                # 如果没有匹配到简单格式，再尝试处理带时间的格式
                                if not record_date:
                                    # 处理带时间的格式：2026-3-25 17:03:44（拼多多格式）
                                    time_formats = [
                                        '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%Y.%m.%d %H:%M:%S',
                                        '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M', '%Y.%m.%d %H:%M'
                                    ]
                                    
                                    for fmt in time_formats:
                                        try:
                                            parsed_datetime = datetime.strptime(date_value, fmt)
                                            # 只提取日期部分，忽略时间部分
                                            record_date = parsed_datetime.strftime('%Y-%m-%d')
                                            print(f"[DEBUG] 时间格式匹配成功：{date_value} -> {record_date}")
                                            break
                                        except:
                                            continue
                                
                                # 如果没有匹配到时间格式，尝试处理单数字月份和日期的格式
                                if not record_date:
                                    # 尝试处理拼多多的单数字格式：2026-3-25 17:03:44
                                    try:
                                        # 使用正则表达式提取日期和时间部分
                                        import re
                                        # 匹配格式：2026-3-25 17:03:44 或 2026/3/25 17:03:44 或 2026.3.25 17:03:44
                                        pattern = r'(\d{4})[-/.]?(\d{1,2})[-/.]?(\d{1,2})\s+(\d{1,2}):(\d{1,2}):(\d{1,2})'
                                        match = re.match(pattern, str(date_value).strip())
                                        if match:
                                            year = int(match.group(1))
                                            month = int(match.group(2))
                                            day = int(match.group(3))
                                            # 验证日期是否有效
                                            if 1 <= month <= 12 and 1 <= day <= 31:
                                                record_date = f"{year:04d}-{month:02d}-{day:02d}"
                                                # 调试信息：显示正则表达式匹配结果
                                                print(f"[DEBUG] 正则匹配成功：{date_value} -> {record_date}")
                                    except Exception as e:
                                        print(f"[DEBUG] 正则匹配错误：{e}")
                                
                                # 如果还是没有匹配到，使用原来的日期解析
                                if not record_date:
                                    record_date = self.parse_date_string(date_value)
                                    print(f"[DEBUG] 使用parse_date_string：{date_value} -> {record_date}")
                                else:
                                    print(f"[DEBUG] 日期识别成功：{date_value} -> {record_date}")
                            elif isinstance(date_value, (datetime,)):
                                # 处理datetime对象，只提取日期部分
                                record_date = date_value.strftime('%Y-%m-%d')
                                print(f"[DEBUG] datetime对象处理：{date_value} -> {record_date}")
                            else:
                                # 其他类型（如Excel日期数字），使用当前日期
                                record_date = self.get_current_date()
                                print(f"[DEBUG] 其他类型，使用当前日期：{date_value} -> {record_date}")
                        
                        # 如果成功识别到日期，就跳出循环
                        if record_date:
                            break
                        else:
                            print(f"[DEBUG] 字段 '{header}' 识别失败，继续尝试其他字段")
                
                # 如果循环结束后还没有识别到日期，检查是否所有字段都尝试过了
                if not record_date:
                    print(f"[DEBUG] 所有日期字段识别失败，使用当前日期")
                
                # 如果没有找到任何日期字段，设置为空（按照用户要求）
                if not record_date:
                    record_date = ''
                    print(f"[DEBUG] Excel表格缺少日期字段，设置为空")
                
                # 添加详细的日期识别调试信息
                print(f"[DEBUG] 最终日期结果：{record_date}")

                # 处理店铺：获取或创建
                store_id = None
                stores = self.db.get_stores()
                for sid, sname in stores:
                    if sname == store_name:
                        store_id = sid
                        break
                if store_id is None:
                    # 自动添加店铺
                    store_id = self.db.add_store(store_name)
                    if store_id is None:
                        fail_count += 1
                        continue
                    self.load_stores()  # 刷新下拉框

                # 检查订单号是否存在
                existing = self.db.get_record_by_order_no(order_no)
                if existing:
                    # 比较数据是否一致
                    same = (existing['store_name'] == store_name and
                            existing['reason'] == reason and
                            abs(existing['refund_amount'] - refund_amount) < 0.01 and
                            existing['cancel'] == cancel and
                            existing['compensate'] == compensate and
                            abs(existing['comp_amount'] - comp_amount) < 0.01 and
                            existing['reject'] == reject and
                            existing['reject_result'] == reject_result and
                            existing['notes'] == notes and
                            existing['record_date'] == record_date)
                    if same:
                        skip_count += 1
                        continue
                    else:
                        # 记录重复订单信息
                        duplicate_orders.append({
                            'order_no': order_no,
                            'existing_data': existing,
                            'new_data': {
                                'store_id': store_id,
                                'order_no': order_no,
                                'reason': reason,
                                'refund_amount': refund_amount,
                                'cancel': cancel,
                                'compensate': compensate,
                                'comp_amount': comp_amount,
                                'reject': reject,
                                'reject_result': reject_result,
                                'notes': notes,
                                'record_date': record_date
                            }
                        })
                else:
                    # 新增订单，直接添加到有效行
                    valid_rows.append({
                        'store_id': store_id,
                        'order_no': order_no,
                        'reason': reason,
                        'refund_amount': refund_amount,
                        'cancel': cancel,
                        'compensate': compensate,
                        'comp_amount': comp_amount,
                        'reject': reject,
                        'reject_result': reject_result,
                        'notes': notes,
                        'record_date': record_date
                    })
            except Exception as e:
                fail_count += 1
                print(f"导入错误：{e}")
        
        # 第二步：如果有重复订单，提供详细处理选项
        if duplicate_orders:
            duplicate_count = len(duplicate_orders)
            
            # 创建详细的选择对话框
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("发现重复订单")
            msg_box.setIcon(QMessageBox.Question)
            
            # 显示详细的重复订单信息（包含店铺名称）
            duplicate_info = f"发现 {duplicate_count} 条重复订单（订单号已存在但数据不一致）\n\n"
            duplicate_info += f"重复订单示例：\n"
            for i, dup in enumerate(duplicate_orders[:5]):
                existing_store = dup['existing_data']['store_name']
                new_store = dup['new_data']['store_name'] if 'store_name' in dup['new_data'] else "导入文件中的店铺"
                duplicate_info += f"{i+1}. 订单号：{dup['order_no']} | 现有店铺：{existing_store} | 导入店铺：{new_store}\n"
            
            if duplicate_count > 5:
                duplicate_info += f"...等{duplicate_count}个订单\n"
            
            duplicate_info += "\n请选择处理方式："
            msg_box.setText(duplicate_info)
            
            # 添加自定义按钮（支持换行）
            overwrite_all_btn = msg_box.addButton("覆盖所有\n重复订单", QMessageBox.YesRole)
            skip_all_btn = msg_box.addButton("跳过所有\n重复订单", QMessageBox.NoRole)
            review_each_btn = msg_box.addButton("逐条查看\n并选择", QMessageBox.ActionRole)
            cancel_btn = msg_box.addButton("取消导入", QMessageBox.RejectRole)
            
            # 设置按钮样式（支持换行和变大）
            for btn in [overwrite_all_btn, skip_all_btn, review_each_btn, cancel_btn]:
                btn.setStyleSheet("""
                    QPushButton {
                        font-size: 12px;
                        padding: 8px 12px;
                        min-height: 40px;
                        min-width: 100px;
                    }
                """)
            
            msg_box.setDefaultButton(overwrite_all_btn)
            msg_box.exec_()
            
            clicked_button = msg_box.clickedButton()
            
            if clicked_button == overwrite_all_btn:
                # 覆盖所有重复订单（处理店铺名称不一致）
                current_search_store = self.search_store_combo.currentText()
                for dup in duplicate_orders:
                    existing_store = dup['existing_data']['store_name']
                    new_store = dup['new_data']['store_name'] if 'store_name' in dup['new_data'] else ""
                    
                    # 如果店铺名称不一致且当前搜索筛选选择了具体店铺，使用当前店铺
                    if existing_store != new_store and current_search_store and current_search_store != "全部":
                        # 获取当前店铺的ID
                        stores = self.db.get_stores()
                        current_store_id = None
                        for sid, sname in stores:
                            if sname == current_search_store:
                                current_store_id = sid
                                break
                        
                        if current_store_id:
                            # 使用当前搜索筛选的店铺
                            dup['new_data']['store_id'] = current_store_id
                            dup['new_data']['store_name'] = current_search_store
                    
                    self.db.update_record(dup['existing_data']['id'], 
                                         dup['new_data']['store_id'],
                                         dup['new_data']['order_no'],
                                         dup['new_data']['reason'],
                                         dup['new_data']['refund_amount'],
                                         dup['new_data']['cancel'],
                                         dup['new_data']['compensate'],
                                         dup['new_data']['comp_amount'],
                                         dup['new_data']['reject'],
                                         dup['new_data']['reject_result'],
                                         dup['new_data']['notes'],
                                         dup['new_data']['record_date'])
                    overwrite_count += 1
                    self.highlighted_orders.add(dup['order_no'])
                    
                    # 立即刷新表格显示，确保状态变化实时显示
                    self.load_table_data(force_reload=True)
            elif clicked_button == skip_all_btn:
                # 跳过所有重复订单
                skip_count += duplicate_count
            elif clicked_button == review_each_btn:
                # 逐条查看重复订单
                for dup in duplicate_orders:
                    existing = dup['existing_data']
                    new_data = dup['new_data']
                    
                    # 显示详细的对比信息（包含店铺名称不一致处理）
                    comparison_info = f"订单号：{dup['order_no']}\n\n"
                    comparison_info += "【现有数据】\n"
                    comparison_info += f"店铺：{existing['store_name']}\n"
                    comparison_info += f"退款原因：{existing['reason']}\n"
                    comparison_info += f"退款金额：¥{existing['refund_amount']}\n"
                    comparison_info += f"登记日期：{existing['record_date']}\n\n"
                    
                    comparison_info += "【导入数据】\n"
                    comparison_info += f"店铺：{new_data['store_name']}\n"
                    comparison_info += f"退款原因：{new_data['reason']}\n"
                    comparison_info += f"退款金额：¥{new_data['refund_amount']}\n"
                    comparison_info += f"登记日期：{new_data['record_date']}\n\n"
                    
                    # 添加店铺名称不一致提示
                    if existing['store_name'] != new_data['store_name']:
                        current_search_store = self.search_store_combo.currentText()
                        if current_search_store and current_search_store != "全部":
                            comparison_info += f"⚠️ 店铺名称不一致，将使用当前筛选的店铺：{current_search_store}\n\n"
                        else:
                            comparison_info += f"⚠️ 店铺名称不一致：现有({existing['store_name']}) vs 导入({new_data['store_name']})\n\n"
                    
                    comparison_info += "请选择处理方式："
                    
                    review_msg_box = QMessageBox(self)
                    review_msg_box.setWindowTitle("重复订单处理")
                    review_msg_box.setIcon(QMessageBox.Question)
                    review_msg_box.setText(comparison_info)
                    
                    overwrite_btn = review_msg_box.addButton("覆盖现有\n数据", QMessageBox.YesRole)
                    skip_btn = review_msg_box.addButton("跳过此\n订单", QMessageBox.NoRole)
                    review_msg_box.addButton("取消剩余\n导入", QMessageBox.RejectRole)
                    
                    # 设置按钮样式
                    for btn in [overwrite_btn, skip_btn]:
                        btn.setStyleSheet("""
                            QPushButton {
                                font-size: 12px;
                                padding: 8px 12px;
                                min-height: 40px;
                                min-width: 80px;
                            }
                        """)
                    
                    review_msg_box.setDefaultButton(overwrite_btn)
                    review_msg_box.exec_()
                    
                    clicked_review_button = review_msg_box.clickedButton()
                    
                    if clicked_review_button == overwrite_btn:
                        # 覆盖此订单（处理店铺名称不一致）
                        current_search_store = self.search_store_combo.currentText()
                        if existing['store_name'] != new_data['store_name'] and current_search_store and current_search_store != "全部":
                            # 获取当前店铺的ID
                            stores = self.db.get_stores()
                            current_store_id = None
                            for sid, sname in stores:
                                if sname == current_search_store:
                                    current_store_id = sid
                                    break
                            
                            if current_store_id:
                                # 使用当前搜索筛选的店铺
                                new_data['store_id'] = current_store_id
                                new_data['store_name'] = current_search_store
                        
                        self.db.update_record(existing['id'], 
                                             new_data['store_id'],
                                             new_data['order_no'],
                                             new_data['reason'],
                                             new_data['refund_amount'],
                                             new_data['cancel'],
                                             new_data['compensate'],
                                             new_data['comp_amount'],
                                             new_data['reject'],
                                             new_data['reject_result'],
                                             new_data['notes'],
                                             new_data['record_date'])
                        overwrite_count += 1
                        self.highlighted_orders.add(dup['order_no'])
                        
                        # 立即刷新表格显示，确保状态变化实时显示
                        self.load_table_data(force_reload=True)
                    elif clicked_review_button == skip_btn:
                        # 跳过此订单
                        skip_count += 1
                    else:
                        # 取消剩余导入
                        skip_count += len(duplicate_orders) - duplicate_orders.index(dup) - 1
                        break
            else:
                # 取消导入
                QMessageBox.information(self, "导入取消", "导入操作已取消")
                return
        
        # 第三步：处理新增订单
        for row_data in valid_rows:
            try:
                self.db.add_record(row_data['store_id'],
                                  row_data['order_no'],
                                  row_data['reason'],
                                  row_data['refund_amount'],
                                  row_data['cancel'],
                                  row_data['compensate'],
                                  row_data['comp_amount'],
                                  row_data['reject'],
                                  row_data['reject_result'],
                                  row_data['notes'],
                                  row_data['record_date'])
                success_count += 1
                self.highlighted_orders.add(row_data['order_no'])
            except Exception as e:
                fail_count += 1
                print(f"新增订单错误：{e}")

        # 显示详细的导入结果
        total_processed = success_count + overwrite_count + skip_count + fail_count
        
        # 创建详细的导入结果对话框
        result_msg = f"导入完成！\n\n"
        result_msg += f"📊 导入统计：\n"
        result_msg += f"• 文件总数据：{total_rows} 条\n"
        result_msg += f"• 成功导入：{success_count} 条\n"
        result_msg += f"• 覆盖重复：{overwrite_count} 条\n"
        result_msg += f"• 跳过重复：{skip_count} 条\n"
        result_msg += f"• 导入失败：{fail_count} 条\n\n"
        
        if duplicate_count > 0:
            result_msg += f"⚠️ 发现重复订单：{duplicate_count} 条\n"
        
        if fail_count > 0:
            result_msg += f"❌ 失败原因：数据格式错误或必填字段缺失\n"
        
        if success_count + overwrite_count > 0:
            result_msg += f"✅ 成功处理：{success_count + overwrite_count} 条数据已保存"
        
        # 显示详细结果对话框
        QMessageBox.information(self, "导入结果", result_msg)
        
        # 同时显示气泡提示
        if success_count == 0 and overwrite_count == 0 and skip_count == 0 and fail_count == 0:
            self.show_tooltip("没有导入数据", "rgba(255, 193, 7, 0.95)", 1500)  # 黄色气泡显示1.5秒
        else:
            self.show_tooltip(f"导入完成 {success_count + overwrite_count}条", "rgba(76, 175, 80, 0.95)", 1500)  # 绿色气泡显示1.5秒
        
        # 调试信息：显示导入的日期范围
        print(f"[DEBUG] 导入完成，成功导入 {success_count} 条记录")
        
        # 智能刷新表格：根据当前筛选条件决定是否显示所有记录
        current_search_store = self.search_store_combo.currentText()
        
        # 如果当前选择了具体店铺，导入的数据应该立即显示
        if current_search_store and current_search_store != "全部":
            # 保持当前筛选条件，但强制刷新表格
            self.load_table_data(force_reload=True)
        else:
            # 没有选择具体店铺，导入后显示所有记录
            self.load_table_data(force_reload=True)
        
        # 强制刷新表格显示
        self.table.viewport().update()
        
        # 检查导入的记录是否显示
        displayed_count = self.table.rowCount()
        imported_count = success_count + overwrite_count
        
        # 如果导入的记录没有显示，提示用户
        if imported_count > 0 and displayed_count == 0:
            QMessageBox.information(self, "导入提示", 
                                  f"✅ 成功导入 {imported_count} 条记录！\n"
                                  f"但当前筛选条件下没有显示任何记录。\n"
                                  f"建议点击【显示全部】按钮查看所有记录。")
        elif imported_count > 0:
            QMessageBox.information(self, "导入成功", 
                                  f"✅ 成功导入 {imported_count} 条记录！\n"
                                  f"当前显示 {displayed_count} 条记录。")
        
        # 设置一个定时器，在用户点击表格后清除高亮（在on_item_clicked中处理）

    def _check_store_exists(self, store_name):
        """检查店铺名称是否存在"""
        try:
            if not self.db or not self.db.conn:
                return False
            cursor = self.db.conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM stores WHERE store_name = ?', (store_name,))
            result = cursor.fetchone()
            if result and isinstance(result, (tuple, list)) and len(result) > 0:
                return result[0] > 0
            return False
        except:
            return False

    def _check_reason_exists(self, reason):
        """检查退款原因是否存在"""
        try:
            if not self.db or not self.db.conn:
                return False
            cursor = self.db.conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM refund_records WHERE reason = ?', (reason,))
            result = cursor.fetchone()
            if result and isinstance(result, (tuple, list)) and len(result) > 0:
                return result[0] > 0
            return False
        except:
            return False

    def check_data_consistency(self):
        """检查数据库和本地表格数据一致性（比较总记录数）"""
        try:
            # 确保数据库连接正常
            if not self.db or not self.db.conn:
                QMessageBox.critical(self, "数据库错误", "数据库连接异常，请重启程序")
                return
            
            # 获取数据库总记录数（所有记录，不考虑筛选条件）
            total_db_count = self.db.get_total_record_count()
            
            # 获取本地表格显示的总行数（当前显示的所有记录）
            local_count = self.table.rowCount() if hasattr(self, 'table') else 0
            
            # 显示核对结果（简化显示，只显示总条数）
            result_msg = f"📊 数据核对结果\n\n"
            result_msg += f"• 数据库总记录数：{total_db_count} 条\n"
            result_msg += f"• 本地表格总记录数：{local_count} 条\n\n"
            
            if total_db_count == local_count:
                result_msg += "✅ 数据一致！数据库和本地表格记录数匹配。"
                QMessageBox.information(self, "数据核对", result_msg)
            else:
                result_msg += f"⚠️ 数据不一致！相差 {abs(total_db_count - local_count)} 条记录。\n\n"
                
                if total_db_count > local_count:
                    result_msg += f"数据库中有 {total_db_count - local_count} 条记录未在本地显示。\n"
                    result_msg += "可能原因：数据缓存问题或筛选条件导致记录被隐藏。"
                else:
                    result_msg += f"本地表格显示 {local_count - total_db_count} 条记录在数据库中不存在。\n"
                    result_msg += "可能原因：数据未保存或数据库连接问题。"
                
                # 提供同步选项
                msg_box = QMessageBox(self)
                msg_box.setWindowTitle("数据不一致")
                msg_box.setIcon(QMessageBox.Warning)
                msg_box.setText(result_msg)
                
                # 添加自定义按钮（支持换行的大按钮）
                sync_btn = msg_box.addButton("同步到本地表格\n（清除所有筛选）", QMessageBox.YesRole)
                sync_btn.setMinimumWidth(180)  # 设置按钮最小宽度
                
                force_sync_btn = msg_box.addButton("强制全局同步\n（清理所有不一致）", QMessageBox.ActionRole)
                force_sync_btn.setMinimumWidth(180)
                
                cleanup_btn = msg_box.addButton("清理数据库\n孤儿记录", QMessageBox.ActionRole)
                cleanup_btn.setMinimumWidth(180)
                
                refresh_btn = msg_box.addButton("刷新表格", QMessageBox.NoRole)
                refresh_btn.setMinimumWidth(120)
                
                cancel_btn = msg_box.addButton("取消", QMessageBox.RejectRole)
                cancel_btn.setMinimumWidth(120)
                
                msg_box.setDefaultButton(sync_btn)
                msg_box.exec_()
                
                clicked_button = msg_box.clickedButton()
                
                if clicked_button == sync_btn:
                    # 强制重新加载表格数据，清除所有筛选条件
                    if hasattr(self, '_cached_records'):
                        self._cached_records = None  # 清除缓存
                    if hasattr(self, '_last_search_params'):
                        self._last_search_params = None  # 清除搜索参数缓存
                    
                    # 清除所有筛选条件
                    self.search_order_edit.clear()
                    self.search_reason_combo.clearChecked()
                    self.search_cancel_combo.setCurrentText('全部')
                    self.search_compensate_combo.setCurrentText('全部')
                    self.search_reject_combo.setCurrentText('全部')
                    self.search_reject_result_combo.setCurrentText('全部')
                    self.search_store_combo.setCurrentText('全部')
                    
                    # 强制重新加载所有数据（从数据库下载到本地）
                    if hasattr(self, 'load_table_data'):
                        self.load_table_data(force_reload=True)
                    
                    # 重新检查一致性
                    new_local_count = self.table.rowCount() if hasattr(self, 'table') else 0
                    if new_local_count == total_db_count:
                        QMessageBox.information(self, "同步成功", 
                                               f"✅ 数据同步完成！\n\n"
                                               f"数据库数据已下载到本地表格。\n"
                                               f"本地表格现在显示 {new_local_count} 条记录，与数据库一致。")
                    else:
                        # 如果仍然不一致，显示调试信息
                        debug_records = self.db.debug_database_records()
                        debug_info = f"数据库中有 {len(debug_records)} 条记录：\n"
                        for record in debug_records:
                            debug_info += f"ID:{record['id']} 订单:{record['order_no']} 店铺:{record['store_name']}\n"
                        
                        QMessageBox.warning(self, "同步失败", 
                                           f"同步后仍然不一致。\n"
                                           f"数据库：{total_db_count}条，本地：{new_local_count}条\n\n"
                                           f"调试信息：\n{debug_info}")
                elif clicked_button == force_sync_btn:
                    # 强制全局同步：彻底清理所有不一致数据
                    sync_result = self.db.force_global_sync()
                    
                    # 清除所有筛选条件并刷新表格
                    if hasattr(self, '_cached_records'):
                        self._cached_records = None
                    if hasattr(self, '_last_search_params'):
                        self._last_search_params = None
                    
                    self.search_order_edit.clear()
                    self.search_reason_combo.clearChecked()
                    self.search_cancel_combo.setCurrentText('全部')
                    self.search_compensate_combo.setCurrentText('全部')
                    self.search_reject_combo.setCurrentText('全部')
                    self.search_reject_result_combo.setCurrentText('全部')
                    self.search_store_combo.setCurrentText('全部')
                    
                    if hasattr(self, 'load_table_data'):
                        self.load_table_data()
                    
                    # 重新检查一致性
                    new_total_db_count = self.db.get_total_record_count()
                    new_local_count = self.table.rowCount() if hasattr(self, 'table') else 0
                    
                    if sync_result['total_cleaned'] > 0:
                        QMessageBox.information(self, "强制同步完成", 
                                               f"✅ 强制全局同步完成！\n\n"
                                               f"清理统计：\n"
                                               f"• 孤儿记录：{sync_result['orphan_count']} 条\n"
                                               f"• 重复记录：{sync_result['duplicate_count']} 条\n"
                                               f"• 无效数据：{sync_result['invalid_count']} 条\n"
                                               f"• 总计清理：{sync_result['total_cleaned']} 条\n\n"
                                               f"同步后：\n"
                                               f"• 数据库总记录数：{new_total_db_count} 条\n"
                                               f"• 本地表格显示记录数：{new_local_count} 条")
                    else:
                        QMessageBox.information(self, "无需清理", "数据库中没有发现不一致数据。")
                elif clicked_button == cleanup_btn:
                    # 清理数据库孤儿记录
                    deleted_count = self.db.cleanup_orphan_records()
                    
                    # 重新检查一致性
                    new_total_db_count = self.db.get_total_record_count()
                    new_local_count = self.table.rowCount() if hasattr(self, 'table') else 0
                    
                    if deleted_count > 0:
                        QMessageBox.information(self, "清理完成", 
                                               f"成功清理 {deleted_count} 条孤儿记录！\n\n"
                                               f"清理后：\n"
                                               f"• 数据库总记录数：{new_total_db_count} 条\n"
                                               f"• 本地表格显示记录数：{new_local_count} 条")
                    else:
                        QMessageBox.information(self, "无需清理", "数据库中没有发现孤儿记录。")
                elif clicked_button == refresh_btn:
                    # 简单刷新表格
                    if hasattr(self, 'load_table_data'):
                        self.load_table_data()
                    QMessageBox.information(self, "刷新完成", "表格已刷新")
        
        except Exception as e:
            # 更详细的错误信息
            import traceback
            error_details = traceback.format_exc()
            QMessageBox.critical(self, "核对错误", 
                               f"数据核对过程中发生错误：{str(e)}\n\n错误详情：\n{error_details}")

    def show_theme_settings(self):
        """显示主题设置对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("主题设置")
        dialog.setFixedSize(400, 400)
        
        layout = QVBoxLayout(dialog)
        
        # 字体颜色设置
        layout.addWidget(QLabel("字体颜色设置："))
        
        # 字体颜色预览
        font_color_layout = QHBoxLayout()
        font_color_layout.addWidget(QLabel("当前字体颜色："))
        self.font_color_preview = QLabel("示例文本")
        self.font_color_preview.setFixedSize(80, 30)
        self.font_color_preview.setStyleSheet("color: black; border: 1px solid black; padding: 5px;")
        font_color_layout.addWidget(self.font_color_preview)
        
        # 字体颜色选择按钮
        self.font_color_picker_btn = QPushButton("选择字体颜色")
        self.font_color_picker_btn.clicked.connect(self.pick_font_color)
        font_color_layout.addWidget(self.font_color_picker_btn)
        
        layout.addLayout(font_color_layout)
        
        # 选中行颜色设置
        layout.addWidget(QLabel("\n选中行颜色设置："))
        
        # 当前颜色预览
        color_layout = QHBoxLayout()
        color_layout.addWidget(QLabel("当前选中行颜色："))
        self.color_preview = QLabel()
        self.color_preview.setFixedSize(50, 30)
        self.color_preview.setStyleSheet("background-color: #87CEEB; border: 1px solid black;")
        color_layout.addWidget(self.color_preview)
        
        # 颜色选择按钮
        self.color_picker_btn = QPushButton("选择颜色")
        self.color_picker_btn.clicked.connect(self.pick_selection_color)
        color_layout.addWidget(self.color_picker_btn)
        
        layout.addLayout(color_layout)
        
        # 店铺颜色管理
        layout.addWidget(QLabel("\n店铺颜色管理："))
        
        # 店铺颜色列表
        self.store_color_list = QListWidget()
        layout.addWidget(self.store_color_list)
        
        # 店铺颜色操作按钮
        store_color_layout = QHBoxLayout()
        self.set_store_color_btn = QPushButton("设置店铺颜色")
        self.set_store_color_btn.clicked.connect(self.set_store_color)
        store_color_layout.addWidget(self.set_store_color_btn)
        
        self.clear_store_color_btn = QPushButton("清除店铺颜色")
        self.clear_store_color_btn.clicked.connect(self.clear_store_color)
        store_color_layout.addWidget(self.clear_store_color_btn)
        
        layout.addLayout(store_color_layout)
        
        # 确定按钮
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(dialog.accept)
        layout.addWidget(ok_btn)
        
        # 加载店铺颜色列表
        self.load_store_colors()
        
        dialog.exec_()

    def pick_font_color(self):
        """选择字体颜色"""
        color = QColorDialog.getColor(QColor("black"), self, "选择字体颜色")
        if color.isValid():
            # 更新字体颜色预览
            self.font_color_preview.setStyleSheet(f"color: {color.name()}; border: 1px solid black; padding: 5px;")
            # 更新表格字体颜色
            self.table.setStyleSheet(f"""
                background-color: white;
                selection-background-color: #87CEEB;  /* 选中背景 */
                selection-color: {color.name()};  /* 自定义选中文字颜色 */
                color: {color.name()};  /* 自定义表格字体颜色 */
            """)

    def pick_selection_color(self):
        """选择选中行颜色"""
        color = QColorDialog.getColor(QColor("#87CEEB"), self, "选择选中行颜色")
        if color.isValid():
            # 更新预览
            self.color_preview.setStyleSheet(f"background-color: {color.name()}; border: 1px solid black;")
            # 更新表格选中颜色
            self.table.setStyleSheet(f"""
                background-color: white;
                selection-background-color: {color.name()};  /* 自定义选中背景 */
                selection-color: black;  /* 黑色选中文字 */
            """)

    def load_store_colors(self):
        """加载店铺颜色列表（显示店铺颜色）"""
        self.store_color_list.clear()
        stores = self.db.get_stores()
        for store_id, store_name in stores:
            # 获取店铺颜色
            store_color = self.db.get_store_color(store_name)
            
            item = QListWidgetItem(f"{store_name}")
            
            # 如果店铺有设置颜色，显示对应的背景色
            if store_color:
                item.setBackground(QColor(store_color))
                # 根据背景色深浅调整文字颜色，确保可读性
                color = QColor(store_color)
                if color.lightness() > 128:  # 浅色背景用黑色文字
                    item.setForeground(QColor("black"))
                else:  # 深色背景用白色文字
                    item.setForeground(QColor("white"))
            
            self.store_color_list.addItem(item)

    def set_store_color(self):
        """设置店铺颜色"""
        current_item = self.store_color_list.currentItem()
        if current_item:
            store_name = current_item.text()
            color = QColorDialog.getColor(QColor("#FFFFFF"), self, f"选择 {store_name} 的颜色")
            if color.isValid():
                # 保存店铺颜色到数据库
                if self.db.set_store_color(store_name, color.name()):
                    QMessageBox.information(self, "提示", f"已为店铺 {store_name} 设置颜色：{color.name()}")
                    # 刷新表格以显示新颜色
                    self.load_table_data()
                    # 刷新店铺列表以显示新颜色
                    self.load_store_colors()

    def clear_store_color(self):
        """清除店铺颜色"""
        current_item = self.store_color_list.currentItem()
        if current_item:
            store_name = current_item.text()
            # 清除店铺颜色
            if self.db.clear_store_color(store_name):
                QMessageBox.information(self, "提示", f"已清除店铺 {store_name} 的颜色设置")
                # 刷新表格以清除颜色
                self.load_table_data()
                # 刷新店铺列表以清除颜色显示
                self.load_store_colors()






                self.db.update_record(record['id'], record['store_id'], record['order_no'], 
                                     record['reason'], record['refund_amount'], 
                                     record['cancel'], record['compensate'], record['comp_amount'],
                                     new_value == "是", reject_result, record['notes'], 
                                     record['record_date'])
        
        # 使用activated信号而不是currentTextChanged，避免频繁触发
        combo.activated.connect(lambda index: on_selection_changed(combo.itemText(index)))

    def show_reject_result_dropdown(self, row, column):
        """显示驳回结果列下拉框选择"""
        # 创建下拉框
        combo = QComboBox()
        combo.addItems(["成功", "失败"])
        
        # 设置当前值
        current_text = self.table.item(row, column).text()
        current_index = combo.findText(current_text)
        if current_index >= 0:
            combo.setCurrentIndex(current_index)
        
        # 显示下拉框
        self.table.setCellWidget(row, column, combo)
        combo.showPopup()
        
        # 为下拉框安装事件过滤器，处理点击空白处关闭
        combo.installEventFilter(self)
        
        # 当下拉框选择改变时更新数据
        def on_selection_changed(new_value):
            self.table.removeCellWidget(row, column)
            self.table.item(row, column).setText(new_value)
            # 保持当前行的选中状态，不清除焦点和选中
            # 强制刷新表格，确保样式更新
            self.table.viewport().update()
            
            # 更新数据库
            order_no = self.table.item(row, 1).text()
            record = self.db.get_record_by_order_no(order_no)
            if record:
                self.db.update_record(record['id'], record['store_id'], record['order_no'], 
                                     record['reason'], record['refund_amount'], 
                                     record['cancel'], record['compensate'], record['comp_amount'],
                                     record['reject'], new_value, record['notes'], 
                                     record['record_date'])
        
        # 使用activated信号而不是currentTextChanged，避免频繁触发
        combo.activated.connect(lambda index: on_selection_changed(combo.itemText(index)))

    def eventFilter(self, obj, event):
        """事件过滤器：处理下拉框点击空白处关闭"""
        if isinstance(obj, QComboBox):
            if event.type() == event.MouseButtonPress:
                # 获取全局鼠标位置
                global_pos = event.globalPos()
                # 获取下拉框的全局位置
                combo_global_rect = QRect(obj.mapToGlobal(QPoint(0, 0)), obj.size())
                
                # 检查点击是否在下拉框外部
                if not combo_global_rect.contains(global_pos):
                    # 直接移除当前下拉框
                    for row in range(self.table.rowCount()):
                        for col in range(self.table.columnCount()):
                            if self.table.cellWidget(row, col) == obj:
                                self.table.removeCellWidget(row, col)
                                # 保持当前行的选中状态，不清除焦点和选中
                                # 强制刷新表格，确保样式更新
                                self.table.viewport().update()
                                return True
        return super().eventFilter(obj, event)

    def copy_order_no(self, row):
        """复制订单号到剪贴板并显示提示气泡"""
        try:
            # 获取订单号
            order_item = self.table.item(row, 1)  # 第1列是订单号
            if order_item:
                order_no = order_item.text()
                
                # 复制到剪贴板
                clipboard = QApplication.clipboard()
                clipboard.setText(order_no)
                
                # 显示提示气泡
                self.show_copy_tooltip(order_no)
                
        except Exception as e:
            QMessageBox.warning(self, "复制失败", f"复制订单号失败：{str(e)}")
    
    def show_refresh_tooltip(self):
        """显示刷新成功的丝滑气泡提示"""
        # 创建提示标签
        tooltip = QLabel("已刷新", self)
        tooltip.setStyleSheet("""
            QLabel {
                background-color: rgba(0, 120, 212, 0.95);
                color: white;
                padding: 10px 16px;
                border-radius: 6px;
                font-size: 13px;
                font-weight: bold;
                border: 1px solid rgba(255, 255, 255, 0.3);
            }
        """)
        tooltip.setAlignment(Qt.AlignCenter)
        tooltip.adjustSize()
        
        # 设置位置（在窗口底部中间显示）
        window_width = self.width()
        window_height = self.height()
        tooltip_x = (window_width - tooltip.width()) // 2
        tooltip_y = window_height - tooltip.height() - 50  # 距离底部50像素
        tooltip.move(tooltip_x, tooltip_y)
        
        # 显示提示
        tooltip.show()
        tooltip.raise_()  # 确保在最上层
        
        # 设置淡入淡出动画
        tooltip.setWindowOpacity(0.0)
        
        # 淡入动画（更快更丝滑）
        fade_in = QTimer(self)
        fade_in.setSingleShot(True)
        fade_in.timeout.connect(lambda: self.safe_fade_tooltip(tooltip, 0.0, 1.0, 200))
        fade_in.start(10)
        
        # 0.8秒后淡出并销毁（更短的显示时间）
        fade_out = QTimer(self)
        fade_out.setSingleShot(True)
        fade_out.timeout.connect(lambda: self.safe_fade_tooltip(tooltip, 1.0, 0.0, 200, True))
        fade_out.start(810)

    def show_tooltip(self, message, color="rgba(0, 120, 212, 0.95)", duration=800):
        """显示通用的淡入淡出气泡提示"""
        # 创建提示标签
        tooltip = QLabel(message, self)
        tooltip.setStyleSheet(f"""
            QLabel {{
                background-color: {color};
                color: white;
                padding: 10px 16px;
                border-radius: 6px;
                font-size: 13px;
                font-weight: bold;
                border: 1px solid rgba(255, 255, 255, 0.3);
            }}
        """)
        tooltip.setAlignment(Qt.AlignCenter)
        tooltip.adjustSize()
        
        # 设置位置（在窗口底部中间显示）
        window_width = self.width()
        window_height = self.height()
        tooltip_x = (window_width - tooltip.width()) // 2
        tooltip_y = window_height - tooltip.height() - 50  # 距离底部50像素
        tooltip.move(tooltip_x, tooltip_y)
        
        # 显示提示
        tooltip.show()
        tooltip.raise_()  # 确保在最上层
        
        # 设置淡入淡出动画
        tooltip.setWindowOpacity(0.0)
        
        # 淡入动画（更快更丝滑）
        fade_in = QTimer(self)
        fade_in.setSingleShot(True)
        fade_in.timeout.connect(lambda: self.safe_fade_tooltip(tooltip, 0.0, 1.0, 200))
        fade_in.start(10)
        
        # 指定时间后淡出并销毁
        fade_out = QTimer(self)
        fade_out.setSingleShot(True)
        fade_out.timeout.connect(lambda: self.safe_fade_tooltip(tooltip, 1.0, 0.0, 200, True))
        fade_out.start(duration + 10)

    def show_copy_tooltip(self, order_no):
        """显示复制成功的提示气泡"""
        # 创建提示标签
        tooltip = QLabel("已复制", self)
        tooltip.setStyleSheet("""
            QLabel {
                background-color: rgba(76, 175, 80, 0.95);
                color: white;
                padding: 12px 20px;
                border-radius: 8px;
                font-size: 14px;
                font-weight: bold;
            }
        """)
        tooltip.setAlignment(Qt.AlignCenter)
        tooltip.adjustSize()
        
        # 设置位置（在软件窗口中下部分居中显示）
        window_width = self.width()
        window_height = self.height()
        tooltip_x = (window_width - tooltip.width()) // 2
        tooltip_y = window_height - tooltip.height() - 50  # 距离底部50像素
        tooltip.move(tooltip_x, tooltip_y)
        
        # 显示提示
        tooltip.show()
        
        # 设置淡入淡出动画
        tooltip.setWindowOpacity(0.0)
        
        # 淡入动画
        fade_in = QTimer(self)
        fade_in.setSingleShot(True)
        fade_in.timeout.connect(lambda: self.fade_tooltip(tooltip, 0.0, 1.0, 300))
        fade_in.start(10)
        
        # 1秒后淡出并销毁
        fade_out = QTimer(self)
        fade_out.setSingleShot(True)
        fade_out.timeout.connect(lambda: self.fade_tooltip(tooltip, 1.0, 0.0, 300, True))
        fade_out.start(1100)
    
    def safe_fade_tooltip(self, tooltip, start_opacity, end_opacity, duration, destroy=False):
        """安全的淡入淡出动画效果，检查对象是否仍然存在"""
        # 检查tooltip对象是否仍然存在
        try:
            # 尝试访问tooltip的属性，如果对象已被删除会抛出异常
            if not tooltip or not hasattr(tooltip, 'setWindowOpacity'):
                return
        except RuntimeError:
            # 对象已被删除，直接返回
            return
        
        # 创建定时器实现动画效果
        timer = QTimer(self)
        timer.setInterval(16)  # 约60fps
        
        start_time = datetime.now()
        
        def update_opacity():
            # 每次更新前都检查对象是否仍然存在
            try:
                if not tooltip or not hasattr(tooltip, 'setWindowOpacity'):
                    timer.stop()
                    return
            except RuntimeError:
                timer.stop()
                return
            
            current_time = datetime.now()
            elapsed = (current_time - start_time).total_seconds() * 1000
            
            if elapsed >= duration:
                try:
                    tooltip.setWindowOpacity(end_opacity)
                    timer.stop()
                    if destroy:
                        tooltip.deleteLater()
                except RuntimeError:
                    # 对象已被删除，直接停止定时器
                    timer.stop()
                return
            
            # 计算当前透明度
            progress = elapsed / duration
            current_opacity = start_opacity + (end_opacity - start_opacity) * progress
            
            try:
                tooltip.setWindowOpacity(current_opacity)
            except RuntimeError:
                # 对象已被删除，直接停止定时器
                timer.stop()
        
        timer.timeout.connect(update_opacity)
        timer.start()

    def fade_tooltip(self, tooltip, start_opacity, end_opacity, duration, destroy=False):
        """淡入淡出动画效果"""
        # 创建定时器实现动画效果
        timer = QTimer(self)
        timer.setInterval(16)  # 约60fps
        
        start_time = datetime.now()
        
        def update_opacity():
            current_time = datetime.now()
            elapsed = (current_time - start_time).total_seconds() * 1000
            
            if elapsed >= duration:
                tooltip.setWindowOpacity(end_opacity)
                timer.stop()
                if destroy:
                    tooltip.deleteLater()
                return
            
            # 计算当前透明度
            progress = elapsed / duration
            current_opacity = start_opacity + (end_opacity - start_opacity) * progress
            tooltip.setWindowOpacity(current_opacity)
        
        timer.timeout.connect(update_opacity)
        timer.start()

    def get_record_id_by_order_no(self, order_no):
        """根据订单号获取记录ID"""
        record = self.db.get_record_by_order_no(order_no)
        return record['id'] if record else None

    def closeEvent(self, event):
        """关闭窗口时关闭数据库连接"""
        self.db.close()
        event.accept()

# ---------------------------- 高级主题设置对话框 --------------------------------
# ---------------------------- 主程序入口 ---------------------------------
if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = RefundManager()
    window.show()
    sys.exit(app.exec_())